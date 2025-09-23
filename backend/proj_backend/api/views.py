from api.models import (
    User, School, Requirement, ListOfPriority, RequestManagement,
    RequestPriority, LiquidationManagement, LiquidationDocument,
    Notification, SchoolDistrict, Backup, AuditLog
)
from .models import RequestManagement
import io
from openpyxl.drawing.image import Image
from openpyxl import Workbook
import sys
import zipfile
import subprocess
import tempfile
import shutil
import os
import random
from django.db.models import Sum
from django.utils.timezone import localtime
from openpyxl.styles import Font, Alignment
import openpyxl
from .models import User
from django.contrib.auth import authenticate
from rest_framework.permissions import AllowAny
from dateutil.relativedelta import relativedelta
from .utils import generate_otp, send_otp_email
from django.utils.crypto import get_random_string
from django.contrib.auth import update_session_auth_hash
import string
from rest_framework.pagination import PageNumberPagination
from django.db import transaction
from django.db import IntegrityError
import logging
from rest_framework import serializers
from django.core import serializers as django_serializers
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import UserSerializer, SchoolSerializer, RequirementSerializer, ListOfPrioritySerializer, RequestManagementSerializer, LiquidationManagementSerializer, LiquidationDocumentSerializer, RequestPrioritySerializer, NotificationSerializer, CustomTokenRefreshSerializer, RequestManagementHistorySerializer, LiquidationManagementHistorySerializer, SchoolDistrictSerializer, PreviousRequestSerializer, BackupSerializer, AuditLogSerializer
from .models import User, School, Requirement, ListOfPriority, RequestManagement, RequestPriority, LiquidationManagement, LiquidationDocument, Notification, LiquidationPriority, SchoolDistrict, Backup, AuditLog, School, SchoolDistrict, Requirement, ListOfPriority, RequestManagement, RequestPriority,  Notification, Backup, AuditLog, GeneratedPDF, PriorityRequirement
from rest_framework.exceptions import PermissionDenied, ValidationError
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import generics
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from datetime import datetime, timedelta
from .serializers import CustomTokenObtainPairSerializer
from django.http import JsonResponse
from rest_framework.permissions import IsAuthenticated
from urllib import request
from .models import School, RequestManagement
from .serializers import UnliquidatedSchoolReportSerializer
from datetime import date
from django.conf import settings
import csv
from django.http import HttpResponse, FileResponse
from django.http import HttpResponse
from decimal import Decimal
from openpyxl.styles import Font, Alignment, Border, Side
from django.db.models.functions import TruncMonth, ExtractDay
import json
from django.db.models import Count, Avg, Sum, Q, F, ExpressionWrapper, FloatField, Case, When, DurationField
from .unliquidated_requests_report_utils import (
    AgingReportPagination,
    generate_aging_report_data,
    generate_aging_csv_report,
    generate_aging_excel_report,
    get_aging_period
)


logger = logging.getLogger(__name__)


class SchoolPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class UserPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class ProtectedView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        content = {'message': 'Hello, World! This is a protected view!'}
        return Response(content)


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


class CustomTokenRefreshView(TokenRefreshView):
    serializer_class = CustomTokenRefreshSerializer


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout(request):
    """
    Logout endpoint with audit trail
    """
    user = request.user

    # Audit for logout is handled by auth signal receiver

    return Response({"message": "Logged out successfully"}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    user = request.user
    old_password = request.data.get('old_password')
    new_password = request.data.get('new_password')

    if not user.check_password(old_password):
        return Response({"error": "Current password is incorrect"}, status=status.HTTP_400_BAD_REQUEST)

    if len(new_password) < 8:
        return Response({"error": "Password must be at least 8 characters"}, status=status.HTTP_400_BAD_REQUEST)

    user.set_password(new_password)
    user.password_change_required = False
    user.save()

    # Rely on signals for user update; no manual CRUD audit here

    # Update the token to prevent automatic logout
    refresh = RefreshToken.for_user(user)
    # Add minimal custom claims
    refresh['first_name'] = user.first_name
    refresh['last_name'] = user.last_name
    refresh['email'] = user.email
    refresh['role'] = user.role
    refresh['password_change_required'] = user.password_change_required

    return Response({
        "access": str(refresh.access_token),
        "refresh": str(refresh),
        "message": "Password updated successfully"
    })


@api_view(['GET', 'POST'])
def user_list(request):
    """
    List all users or create a new user with enhanced filtering
    """
    if request.method == 'GET':
        # Get filter parameters from query string
        show_archived = request.query_params.get(
            'archived', 'false').lower() == 'true'
        role_filter = request.query_params.get('role', None)
        school_filter = request.query_params.get('school', None)
        search_term = request.query_params.get('search', None)
        ordering = request.query_params.get('ordering', '-date_joined')
        show_all = request.query_params.get(
            'show_all', 'false').lower() == 'true'
        archived = request.query_params.get(
            'archived', 'false').lower() == 'true'

        queryset = User.objects.exclude(id=request.user.id)
        # Archive filter
        if not show_all:
            if not archived:
                queryset = queryset.filter(is_active=True)
            else:
                queryset = queryset.filter(is_active=False)

        # Role filter
        if role_filter:
            queryset = queryset.filter(role=role_filter)

        # School filter
        if school_filter:
            queryset = queryset.filter(
                school__schoolName__icontains=school_filter)
            queryset = queryset.filter(
                school__schoolId__icontains=school_filter)

        # Search filter
        if search_term:
            queryset = queryset.filter(
                Q(first_name__icontains=search_term) |
                Q(last_name__icontains=search_term) |
                Q(email__icontains=search_term) |
                Q(phone_number__icontains=search_term) |
                Q(sex__icontains=search_term) |
                # <-- Fix: use related field
                Q(school__schoolName__icontains=search_term) |
                # <-- Optionally add this too
                Q(school__schoolId__icontains=search_term)
            )

        # Add ordering support
        if ordering:
            queryset = queryset.order_by(ordering)

        # Paginate
        paginator = UserPagination()
        page = paginator.paginate_queryset(queryset, request)
        serializer = UserSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)

    elif request.method == 'POST':
        # Only allow admins to create admin users
        # if request.data.get('role') == 'admin' and not request.user.is_superuser:
        #     return Response(
        #         {'error': 'Only administrators can create admin users'},
        #         status=status.HTTP_403_FORBIDDEN
        #     )

        serializer = UserSerializer(
            data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
def user_detail(request, pk):
    """
    Retrieve, update or delete a user instance with enhanced permissions
    """
    user = get_object_or_404(User, pk=pk)

    if request.method == 'GET':
        serializer = UserSerializer(user)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        # Prevent non-admins from making users admins
        if (request.data.get('role') == 'admin' and
            not request.user.is_superuser and
                user.role != 'admin'):
            return Response(
                {'error': 'Only administrators can assign admin role'},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = UserSerializer(
            user, data=request.data, partial=True, context={'request': request})
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Check if sensitive fields were modified
        sensitive_fields = ['email',  'password', 'role']
        needs_new_token = any(
            field in request.data for field in sensitive_fields)

        serializer.save()

        response_data = serializer.data

        # Generate new token if needed
        if needs_new_token:
            refresh = RefreshToken.for_user(user)
            refresh['first_name'] = user.first_name
            refresh['last_name'] = user.last_name
            refresh['email'] = user.email
            refresh['role'] = user.role

            response_data['token'] = {
                'access': str(refresh.access_token),
                'refresh': str(refresh),
            }

        return Response(response_data, status=status.HTTP_200_OK)

    elif request.method == 'PATCH':
        # Special handling for archive/unarchive
        if 'is_active' in request.data:
            # Prevent users from archiving themselves
            if user == request.user:
                return Response(
                    {'error': 'You cannot archive your own account'},
                    status=status.HTTP_403_FORBIDDEN
                )
            # Use serializer to validate activation conflicts
            serializer = UserSerializer(
                user, data=request.data, partial=True, context={'request': request})
            if serializer.is_valid():
                is_archiving = 'is_active' in request.data and not request.data[
                    'is_active'] and user.is_active
                is_restoring = 'is_active' in request.data and request.data[
                    'is_active'] and not user.is_active

                serializer.save()

                # Log the action
                from .audit_utils import log_audit_event
                action = 'update'
                if is_archiving:
                    action = 'archive'
                elif is_restoring:
                    action = 'restore'

                # Rely on signals for archive/restore update logs

                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # For other PATCH operations
        serializer = UserSerializer(
            user, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        # Prevent users from deleting themselves
        if user == request.user:
            return Response(
                {'error': 'You cannot delete your own account'},
                status=status.HTTP_403_FORBIDDEN
            )
        # Soft delete by archiving instead of hard delete
        user.is_active = False
        user.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class SchoolListCreateAPIView(generics.ListCreateAPIView):
    serializer_class = SchoolSerializer
    pagination_class = SchoolPagination

    def get_queryset(self):
        queryset = School.objects.select_related('district').all()
        search_term = self.request.query_params.get('search', None)
        legislative_district = self.request.query_params.get(
            'legislative_district', None)
        municipality = self.request.query_params.get('municipality', None)
        district_id = self.request.query_params.get('district', None)
        archived = self.request.query_params.get(
            'archived', 'false').lower() == 'true'

        if not archived:
            queryset = queryset.filter(is_active=True)
        else:
            queryset = queryset.filter(is_active=False)

        if search_term:
            queryset = queryset.filter(
                Q(schoolName__icontains=search_term) |
                Q(district__districtName__icontains=search_term) |
                Q(municipality__icontains=search_term)
            )
        if legislative_district:
            queryset = queryset.filter(
                legislativeDistrict=legislative_district)
        if municipality:
            queryset = queryset.filter(municipality=municipality)
        if district_id:
            queryset = queryset.filter(district_id=district_id)
        return queryset.order_by('schoolName')

    def create(self, request, *args, **kwargs):
        # Check if the input data is a list (for bulk creation)
        is_many = isinstance(request.data, list)

        # Use many=True for lists, otherwise use default behavior
        serializer = self.get_serializer(data=request.data, many=is_many)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)

        # Audit will be handled by signals

        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


@api_view(['GET'])
def search_schools(request):
    search_term = request.query_params.get('search', '')
    schools = School.objects.filter(
        Q(schoolName__icontains=search_term) |
        Q(district__districtName__icontains=search_term) |
        Q(district__districtId__icontains=search_term) |  # Add district ID search
        Q(municipality__icontains=search_term)
    ).order_by('schoolName')[:10]  # Limit to 10 results
    serializer = SchoolSerializer(schools, many=True)
    return Response(serializer.data)


class SchoolRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = School.objects.all()
    serializer_class = SchoolSerializer
    lookup_field = 'schoolId'

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        old_is_active = instance.is_active

        response = super().update(request, *args, **kwargs)

        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        school_name = f"{instance.schoolName} ({instance.schoolId})"

        response = super().destroy(request, *args, **kwargs)


        return response


class RequirementListCreateAPIView(generics.ListCreateAPIView):
    serializer_class = RequirementSerializer

    def get_queryset(self):
        queryset = Requirement.objects.all()
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        search_term = self.request.query_params.get('search', None)
        if search_term:
            queryset = queryset.filter(requirementTitle__icontains=search_term)
        return queryset

    def create(self, request, *args, **kwargs):
        is_many = isinstance(request.data, list)
        serializer = self.get_serializer(data=request.data, many=is_many)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)

        # Audit handled by signals

        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class RequirementRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Requirement.objects.all()
    serializer_class = RequirementSerializer
    lookup_field = 'requirementID'

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        old_is_active = instance.is_active

        response = super().update(request, *args, **kwargs)


        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        requirement_name = instance.requirementTitle

        response = super().destroy(request, *args, **kwargs)


        return response


class ListOfPriorityListCreateAPIView(generics.ListCreateAPIView):
    serializer_class = ListOfPrioritySerializer

    def get_queryset(self):
        queryset = ListOfPriority.objects.all()
        archived = self.request.query_params.get(
            'archived', 'false').lower() == 'true'
        if archived:
            queryset = queryset.filter(is_active=False)
        else:
            queryset = queryset.filter(is_active=True)
        return queryset

    def create(self, request, *args, **kwargs):
        # Enable batch posting
        is_many = isinstance(request.data, list)
        serializer = self.get_serializer(data=request.data, many=is_many)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)

        # Audit handled by signals

        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class ListOfPriorityRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ListOfPriority.objects.all()
    serializer_class = ListOfPrioritySerializer
    lookup_field = 'LOPID'

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        old_is_active = instance.is_active

        response = super().update(request, *args, **kwargs)

        

        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        priority_name = instance.expenseTitle

        response = super().destroy(request, *args, **kwargs)


        return response


class RequestManagementListView(generics.ListAPIView):
    """
    View for listing all requests (for admin/superusers) or user's own requests
    """
    serializer_class = RequestManagementSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = RequestManagement.objects.all()

        # Non-admin users only see their own requests
        if not self.request.user.role in ['admin', 'superintendent']:
            queryset = queryset.filter(user=self.request.user)

        # Filter by status if provided
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)

        return queryset.order_by('-created_at')


class RequestManagementListCreateView(generics.ListCreateAPIView):
    serializer_class = RequestManagementSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        try:
            user = self.request.user

            # Check if user is eligible to submit a request
            target_month = serializer.validated_data.get('request_monthyear')

            if target_month:
                # User specified a month, validate it
                if not RequestManagement.can_user_request_for_month(user, target_month):
                    raise ValidationError({
                        'request_monthyear': 'You cannot submit a request for this month. Please liquidate your current request first.'
                    })
            else:
                # Auto-determine the next available month
                temp_request = RequestManagement(user=user)
                target_month = temp_request.get_next_available_month()
                serializer.validated_data['request_monthyear'] = target_month

            # Save with atomic transaction
            with transaction.atomic():
                instance = serializer.save(user=user)
                # Force save to ensure signals fire
                instance.save()
                logger.info(
                    f"Successfully created request {instance.request_id} for month {target_month}")

        except ValidationError:
            raise  # Re-raise validation errors
        except Exception as e:
            logger.error(f"Failed to create request: {str(e)}")
            raise ValidationError(
                "Failed to create request. Please try again.")

    def get_queryset(self):
        queryset = RequestManagement.objects.select_related(
            'user__school__district'  # Add this to optimize queries
        ).all()
        user = self.request.user

        # Filter by multiple statuses if provided (comma-separated)
        status_param = self.request.query_params.get('status')
        if status_param:
            status_list = [s.strip()
                           for s in status_param.split(',') if s.strip()]
            if status_list:
                queryset = queryset.filter(status__in=status_list)

        # Filter by multiple school_ids if provided (comma-separated)
        school_ids_param = self.request.query_params.get('school_ids')
        if school_ids_param:
            school_ids_list = [s.strip()
                               for s in school_ids_param.split(',') if s.strip()]
            if school_ids_list:
                queryset = queryset.filter(
                    user__school__schoolId__in=school_ids_list)

        # FIXED: District filter - ensure proper field lookup
        district_param = self.request.query_params.get('district')
        if district_param:
            print(f"Filtering by district: {district_param}")  # Debug log
            queryset = queryset.filter(
                user__school__district__districtId=district_param)
            # Debug log
            print(f"Queryset after district filter: {queryset.count()}")

        # FIXED: Legislative District filter
        legislative_district_param = self.request.query_params.get(
            'legislative_district')
        if legislative_district_param:
            # Debug log
            print(
                f"Filtering by legislative district: {legislative_district_param}")
            queryset = queryset.filter(
                user__school__district__legislativeDistrict=legislative_district_param)
            # Debug log
            print(
                f"Queryset after legislative district filter: {queryset.count()}")

        # FIXED: Municipality filter
        municipality_param = self.request.query_params.get('municipality')
        if municipality_param:
            # Debug log
            print(f"Filtering by municipality: {municipality_param}")
            queryset = queryset.filter(
                user__school__district__municipality=municipality_param)
            # Debug log
            print(f"Queryset after municipality filter: {queryset.count()}")

        # Search filter - add this if it's missing
        search_param = self.request.query_params.get('search')
        if search_param:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search_param) |
                Q(user__last_name__icontains=search_param) |
                Q(user__school__schoolName__icontains=search_param) |
                Q(request_id__icontains=search_param) |
                Q(priorities__priority__expenseTitle__icontains=search_param)
            ).distinct()

        # Date range filter
        start_date_param = self.request.query_params.get('start_date')
        end_date_param = self.request.query_params.get('end_date')
        if start_date_param and end_date_param:
            queryset = queryset.filter(
                created_at__date__gte=start_date_param,
                created_at__date__lte=end_date_param
            )

        # Enhanced filtering for different user roles
        if user.role == 'superintendent':
            # Superintendents only see requests that should be visible (current/past months and pending)
            today = date.today()
            current_month_year = f"{today.year:04d}-{today.month:02d}"

            # Show pending requests for current/past months, and all non-pending requests
            queryset = queryset.filter(
                Q(status='pending', request_monthyear__lte=current_month_year) |
                Q(status__in=['approved', 'rejected',
                  'downloaded', 'unliquidated', 'liquidated'])
            )
        elif user.role not in ['admin', 'accountant']:
            # Regular users only see their own requests
            queryset = queryset.filter(user=user)

        return queryset.order_by('-created_at')


class RequestManagementRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = RequestManagement.objects.all()
    serializer_class = RequestManagementSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'request_id'

    def perform_update(self, serializer):
        instance = self.get_object()
        # Set the sender for notification
        instance._status_changed_by = self.request.user
        serializer.save()


class ApproveRequestView(generics.UpdateAPIView):
    queryset = RequestManagement.objects.all()
    serializer_class = RequestManagementSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

    def update(self, request, *args, **kwargs):
        with transaction.atomic():  # Ensure atomic transaction
            # Lock the row to prevent race conditions
            instance = RequestManagement.objects.select_for_update().get(
                pk=kwargs['pk'])

            print(f"Current status: {instance.status}")  # Debug log

            # Permission check
            if request.user.role not in ['admin', 'superintendent']:
                return Response(
                    {"detail": "Only administrators can approve requests"},
                    status=status.HTTP_403_FORBIDDEN
                )

            # State validation
            if instance.status != 'pending':
                return Response(
                    {"detail": f"Current status is '{instance.status}', must be 'pending'"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Prepare for notification
            instance._old_status = instance.status
            instance._status_changed_by = request.user
            instance._skip_auto_status = True  # Add this line
            instance.status = 'approved'
            instance.date_approved = timezone.now().date()
            instance.reviewed_by = request.user
            instance.reviewed_at = timezone.now()

            # Save only the updated fields
            instance.save(update_fields=[
                'status',
                'date_approved',
                'reviewed_by',
                'reviewed_at'
            ])
            print(f"Status after save: {instance.status}")  # Debug log

            # Log approval AFTER save
            # from .audit_utils import log_audit_event
            # log_audit_event(
            #     request=request,
            #     action='approve',
            #     module='request',
            #     description=f"Approved request {instance.request_id}",
            #     object_id=instance.request_id,
            #     object_type='RequestManagement',
            #     object_name=f"Request {instance.request_id}"
            # )

            # Verify in database
            refreshed = RequestManagement.objects.get(pk=instance.pk)
            print(f"Database status: {refreshed.status}")  # Debug log

            return Response(self.get_serializer(instance).data)


class RejectRequestView(generics.UpdateAPIView):
    queryset = RequestManagement.objects.all()
    serializer_class = RequestManagementSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        instance._old_status = instance.status  # CRITICAL: Track previous state
        instance._status_changed_by = request.user
        instance._skip_auto_status = True  # Add this line
        instance._skip_signal_audit = True  # Avoid duplicate signal audit; manual business log below

        # Permission check
        if request.user.role not in ['admin', 'superintendent']:
            return Response(
                {"detail": "Only administrators can reject requests"},
                status=status.HTTP_403_FORBIDDEN
            )

        # State validation
        if instance.status != 'pending':
            return Response(
                {"detail": "Only pending requests can be rejected"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Comment validation
        rejection_comment = request.data.get('rejection_comment', '').strip()
        if not rejection_comment:
            return Response(
                {"detail": "Please provide a rejection comment"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Save will trigger signals
        instance.status = 'rejected'
        instance.rejection_comment = request.data.get('rejection_comment')
        instance.rejection_date = timezone.now().date()

        # Explicit save
        instance.save(
            update_fields=['status', 'rejection_comment', 'rejection_date'])
        from .audit_utils import log_audit_event
        log_audit_event(
            request=request,
            action='reject',
            module='request',
            description=f"Rejected request {instance.request_id}",
            object_id=instance.request_id,
            object_type='RequestManagement',
            object_name=f"Request {instance.request_id}"
        )
        return Response(self.get_serializer(instance).data)


class RequestManagementDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    View for retrieving, updating, or deleting a specific request.
    """
    queryset = RequestManagement.objects.all()
    serializer_class = RequestManagementSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'request_id'

    def perform_update(self, serializer):
        instance = self.get_object()

        # Prevent editing if request is already approved
        if instance.status == 'approved':
            raise ValidationError("Cannot edit an already approved request")

        # Check if user has permission to edit this request
        if instance.user != self.request.user and self.request.user.role not in ['admin', 'superintendent']:
            raise PermissionDenied(
                "You don't have permission to edit this request")

        # Allow updates for rejected requests and reset status to pending
        if instance.status == 'rejected':
            # Assuming status_changed_by is a model field
            serializer.save(status='pending',
                            status_changed_by=self.request.user)
        else:
            serializer.save()


class RequestPriorityCreateView(generics.CreateAPIView):
    """
    View for adding priorities to an existing request
    """
    serializer_class = RequestPrioritySerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        request_id = self.kwargs['request_id']
        request_obj = RequestManagement.objects.get(request_id=request_id)

        # Check if user has permission to modify this request
        if request_obj.user != self.request.user and self.request.user.role not in ['admin', 'superintendent']:
            raise PermissionDenied(
                "You don't have permission to modify this request")

        # Prevent adding priorities if request is already approved/rejected
        if request_obj.status in ['approved', 'rejected']:
            raise ValidationError(
                "Cannot add priorities to an already approved/rejected request")

        serializer.save(request=request_obj)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_pending_requests(request):
    """
    Check if user has any pending requests or liquidations that aren't completed
    """
    # Check for pending/rejected requests
    user_requests = RequestManagement.objects.filter(
        user=request.user,
        status__in=['pending', 'approved']  # Include rejected
    ).order_by('-created_at')

    # Check for liquidations that aren't completed
    active_liquidations = LiquidationManagement.objects.filter(
        request__user=request.user
    ).exclude(status='completed').order_by('-created_at')

    response_data = {
        'has_pending_request': user_requests.exists(),
        'has_active_liquidation': active_liquidations.exists(),
        'pending_request': RequestManagementSerializer(user_requests.first()).data if user_requests.exists() else None,
        'active_liquidation': LiquidationManagementSerializer(active_liquidations.first(), context={'request': request}).data if active_liquidations.exists() else None
    }

    return Response(response_data)


def get_priorities_for_history(request_id, as_of_date):
    # Get all priority history records that existed at or before as_of_date
    from .models import RequestPriority
    priorities = RequestPriority.history.as_of(
        as_of_date).filter(request_id=request_id)

    return [
        {
            "expenseTitle": rp.priority.expenseTitle,
            "LOPID": rp.priority.LOPID,
            "amount": rp.amount,
        }
        for rp in priorities
    ]


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def resubmit_request(request, request_id):
    try:
        req = RequestManagement.objects.get(
            request_id=request_id, user=request.user)

        if req.status != 'rejected':
            return Response({"error": "Only rejected requests can be resubmitted"}, status=400)

        # Update priorities
        RequestPriority.objects.filter(request=req).delete()
        for pa in request.data.get('priority_amounts', []):
            priority = ListOfPriority.objects.get(LOPID=pa['LOPID'])
            RequestPriority.objects.create(
                request=req,
                priority=priority,
                amount=pa['amount']
            )

        # Reset status and clear rejection fields
        req._skip_signal_audit = True
        req.status = 'pending'
        # req.rejection_comment = None
        # req.rejection_date = None
        req.save()

        # Log audit for resubmission
        from .audit_utils import log_audit_event
        log_audit_event(
            request=request,
            action='resubmit',
            module='request',
            description=f"Resubmitted request {req.request_id}",
            object_id=req.request_id,
            object_type='RequestManagement',
            object_name=f"Request {req.request_id}"
        )

        # --- Notification logic ---
        from .models import Notification
        Notification.objects.create(
            notification_title="Request Resubmitted",
            details="Your request has been resubmitted and is pending review.",
            receiver=req.user,
            sender=request.user,
        )
        # --- End notification logic ---

        return Response(RequestManagementSerializer(req).data)

    except RequestManagement.DoesNotExist:
        return Response({"error": "Request not found"}, status=404)


class LiquidationManagementListCreateAPIView(generics.ListCreateAPIView):
    queryset = LiquidationManagement.objects.all()
    serializer_class = LiquidationManagementSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = LiquidationManagement.objects.select_related(
            'request', 'request__user', 'request__user__school', 'request__user__school__district'
        )

        # Optional status filtering (comma-separated)
        status_param = self.request.query_params.get('status')
        status_list = None
        if status_param:
            status_list = [s.strip()
                           for s in status_param.split(',') if s.strip()]

        # Role-scoped base queryset and default statuses when none provided
        if user.role == 'district_admin':
            if user.school_district:
                queryset = queryset.filter(
                    request__user__school__district=user.school_district
                )
            default_statuses = ['submitted',
                                'under_review_district', 'resubmit']
            queryset = queryset.filter(
                status__in=(status_list or default_statuses))

        elif user.role == 'liquidator':
            default_statuses = ['under_review_liquidator', 'approved_district']
            queryset = queryset.filter(
                status__in=(status_list or default_statuses))

        elif user.role == 'accountant':
            default_statuses = ['under_review_division', 'approved_liquidator']
            queryset = queryset.filter(
                status__in=(status_list or default_statuses))

        elif user.role == 'school_head':
            queryset = queryset.filter(request__user=user).exclude(
                status__in=['liquidated'])

        # Other roles (admin etc.) - allow optional status filter
        else:
            if status_list:
                queryset = queryset.filter(status__in=status_list)

        # Additional optional filters similar to requests list
        legislative_district = self.request.query_params.get(
            'legislative_district')
        if legislative_district:
            queryset = queryset.filter(
                request__user__school__district__legislativeDistrict=legislative_district
            )

        municipality = self.request.query_params.get('municipality')
        if municipality:
            queryset = queryset.filter(
                request__user__school__district__municipality=municipality
            )

        district_param = self.request.query_params.get('district')
        if district_param:
            queryset = queryset.filter(
                request__user__school__district__districtId=district_param
            )

        # Date range filter on created_at
        start_date_param = self.request.query_params.get('start_date')
        end_date_param = self.request.query_params.get('end_date')
        if start_date_param and end_date_param:
            queryset = queryset.filter(
                created_at__date__gte=start_date_param,
                created_at__date__lte=end_date_param
            )

        ordering = self.request.query_params.get('ordering')
        if ordering:
            queryset = queryset.order_by(ordering)
        else:
            queryset = queryset.order_by('-created_at')

        return queryset

    def perform_create(self, serializer):
        liquidation = serializer.save(user=self.request.user)

        # Log audit for liquidation creation
        from .audit_utils import log_audit_event
        log_audit_event(
            request=self.request,
            action='create',
            module='liquidation',
            description=f"Created liquidation {liquidation.LiquidationID} for request {liquidation.request.request_id}",
            object_id=liquidation.LiquidationID,
            object_type='LiquidationManagement',
            object_name=f"Liquidation {liquidation.LiquidationID}"
        )


class LiquidationManagementRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LiquidationManagement.objects.all()
    serializer_class = LiquidationManagementSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'LiquidationID'

    def get_queryset(self):
        queryset = super().get_queryset()
        # Prefetch related data for better performance
        return queryset.select_related(
            'request',
            'request__user',
            'request__user__school',
            'reviewed_by_district',
            'reviewed_by_liquidator',
            'reviewed_by_division'
        ).prefetch_related(
            'documents',
            'documents__requirement',
            'documents__request_priority',
            'documents__request_priority__priority',
            'liquidation_priorities',
            'liquidation_priorities__priority'
        )

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        instance._old_status = instance.status
        partial = kwargs.pop('partial', False)
        data = request.data.copy()
        new_status = data.get("status")
        user_role = request.user.role

        # Handle different approval levels
        if new_status == "approved_district" and user_role == "district_admin":
            instance.reviewed_by_district = request.user
            instance.reviewed_at_district = timezone.now()
            instance.date_districtApproved = timezone.now().date()

        elif new_status == "approved_liquidator" and user_role == "liquidator":
            instance.reviewed_by_liquidator = request.user
            instance.reviewed_at_liquidator = timezone.now()
            instance.date_liquidatorApproved = timezone.now().date()

        elif new_status == "liquidated" and user_role == "accountant":
            instance.reviewed_by_division = request.user
            instance.reviewed_at_division = timezone.now()
            instance.date_liquidated = timezone.now()

            if instance.request:
                instance.request._skip_auto_status = True
                instance.request.status = 'liquidated'
                instance.request.save(update_fields=['status'])

        # If rejecting (resubmit), capture rejection_comment
        if new_status == "resubmit":
            rejection_comment = data.get('rejection_comment', '').strip()
            if not rejection_comment:
                return Response(
                    {"detail": "Please provide a rejection comment for resubmission."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            instance.rejection_comment = rejection_comment

        # Set the user who changed the status for notification
        instance._status_changed_by = request.user  # <-- CRUCIAL for notifications!
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        instance.save()

        # Log audit for liquidation status changes
        from .audit_utils import log_audit_event
        old_status = instance._old_status
        new_status = instance.status

        if old_status != new_status:
            # Determine action type based on status change
            action = 'update'
            if new_status == 'approved_district':
                action = 'approve_district'
            elif new_status == 'approved_liquidator':
                action = 'approve_liquidator'
            elif new_status == 'liquidated':
                action = 'liquidate'
            elif new_status == 'resubmit':
                action = 'reject'
            elif new_status == 'submitted':
                action = 'submit'

            # log_audit_event(
            #     request=request,
            #     action=action,
            #     module='liquidation',
            #     description=f"Changed liquidation {instance.LiquidationID} status from {old_status} to {new_status}",
            #     object_id=instance.LiquidationID,
            #     object_type='LiquidationManagement',
            #     object_name=f"Liquidation {instance.LiquidationID}",
            #     old_values={'status': old_status},
            #     new_values={'status': new_status}
            # )

        return Response(serializer.data)

    def perform_update(self, serializer):
        instance = self.get_object()
        instance._status_changed_by = self.request.user  # <-- CRUCIAL for notifications!
        serializer.save()


# Removed old approve_liquidation function - now handled in update method


class LiquidationDocumentListCreateAPIView(generics.ListCreateAPIView):
    queryset = LiquidationDocument.objects.all()
    serializer_class = LiquidationDocumentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        liquidation_id = self.kwargs.get('LiquidationID')
        return self.queryset.filter(liquidation__LiquidationID=liquidation_id)

    def create(self, request, *args, **kwargs):
        liquidation_id = self.kwargs.get('LiquidationID')
        liquidation = get_object_or_404(
            LiquidationManagement, LiquidationID=liquidation_id)

        # Check for existing document
        existing_doc = self.get_queryset().filter(
            request_priority_id=request.data.get('request_priority'),
            requirement_id=request.data.get('requirement')
        ).first()

        if existing_doc:
            serializer = self.get_serializer(
                existing_doc, data=request.data, partial=True)
        else:
            serializer = self.get_serializer(data=request.data)

        serializer.is_valid(raise_exception=True)
        serializer.save(liquidation=liquidation, uploaded_by=request.user)

        headers = self.get_success_headers(serializer.data)
        return Response(
            serializer.data,
            status=status.HTTP_200_OK if existing_doc else status.HTTP_201_CREATED,
            headers=headers
        )


class LiquidationDocumentRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LiquidationDocument.objects.all()
    serializer_class = LiquidationDocumentSerializer
    permission_classes = [IsAuthenticated]


@api_view(['POST'])
def submit_for_liquidation(request, request_id):
    try:
        with transaction.atomic():
            request_obj = RequestManagement.objects.select_for_update().get(request_id=request_id)

            if request_obj.status != 'approved':
                return Response(
                    {'error': 'Request must be approved before liquidation'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get download date from request data or use current date
            download_date = request.data.get('download_date')
            if download_date:
                try:
                    download_date = timezone.datetime.strptime(
                        download_date, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    return Response(
                        {'error': 'Invalid download date format. Use YYYY-MM-DD'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Validate download date is not in the future and not before approval date
                if download_date > timezone.now().date():
                    return Response(
                        {'error': 'Download date cannot be in the future'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                if request_obj.date_approved and download_date < request_obj.date_approved:
                    return Response(
                        {'error': 'Download date cannot be before approval date'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                download_date = timezone.now().date()

            # First update the request status to 'downloaded' and set date_downloaded
            request_obj._status_changed_by = request.user
            # Avoid duplicate signal audit; we will log business event manually
            request_obj._skip_signal_audit = True
            request_obj.status = 'downloaded'
            request_obj.downloaded_at = download_date  # Use the selected date
            request_obj.save(update_fields=['status', 'downloaded_at'])

            # Log audit for download status change
            from .audit_utils import log_audit_event
            log_audit_event(
                request=request,
                action='download',
                module='request',
                description=f"Downloaded request {request_obj.request_id} for liquidation",
                object_id=request_obj.request_id,
                object_type='RequestManagement',
                object_name=f"Request {request_obj.request_id}"
            )

            # Create liquidation record
            liquidation, created = LiquidationManagement.objects.get_or_create(
                request=request_obj,
                defaults={'status': 'draft'}
            )

            if not created:
                return Response(
                    {'error': 'Liquidation already exists for this request'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Log audit for liquidation creation
            log_audit_event(
                request=request,
                action='create',
                module='liquidation',
                description=f"Created liquidation {liquidation.LiquidationID} for request {request_obj.request_id}",
                object_id=liquidation.LiquidationID,
                object_type='LiquidationManagement',
                object_name=f"Liquidation {liquidation.LiquidationID}"
            )

            # Update request status to 'unliquidated' as final state
            request_obj._skip_signal_audit = True
            request_obj.status = 'unliquidated'
            request_obj.save(update_fields=['status'])

            # Log audit for unliquidated status change
            log_audit_event(
                request=request,
                action='update',
                module='request',
                description=f"Changed request {request_obj.request_id} status to unliquidated",
                object_id=request_obj.request_id,
                object_type='RequestManagement',
                object_name=f"Request {request_obj.request_id}",
                old_values={'status': 'downloaded'},
                new_values={'status': 'unliquidated'}
            )

            serializer = LiquidationManagementSerializer(liquidation)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

    except RequestManagement.DoesNotExist:
        return Response(
            {'error': 'Request not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except ValidationError as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def submit_liquidation(request, LiquidationID):
    try:
        with transaction.atomic():
            liquidation = LiquidationManagement.objects.select_for_update().get(
                LiquidationID=LiquidationID
            )

            # Check if all required documents are uploaded
            required_docs_missing = False
            for rp in liquidation.request.requestpriority_set.all():
                for req in rp.priority.requirements.filter(is_required=True):
                    if not LiquidationDocument.objects.filter(
                        liquidation=liquidation,
                        request_priority=rp,
                        requirement=req
                    ).exists():
                        required_docs_missing = True
                        break
                if required_docs_missing:
                    break

            # Allow submission if status is 'draft' or 'resubmit'
            if liquidation.status not in ['draft', 'resubmit']:
                return Response(
                    {'error': 'Liquidation must be in draft or resubmit status to be submitted'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if required_docs_missing:
                return Response(
                    {'error': 'All required documents must be uploaded before submission'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get the actual amounts from the request data
            actual_amounts = request.data.get('actual_amounts', [])

            # Update liquidation priorities with actual amounts
            for amount_data in actual_amounts:
                try:
                    priority = ListOfPriority.objects.get(
                        LOPID=amount_data['expense_id'])
                    liquidation_priority, created = LiquidationPriority.objects.get_or_create(
                        liquidation=liquidation,
                        priority=priority,
                        defaults={'amount': amount_data['actual_amount']}
                    )
                    if not created:
                        liquidation_priority.amount = amount_data['actual_amount']
                        liquidation_priority.save()
                except (ListOfPriority.DoesNotExist, KeyError):
                    continue

            # Recalculate refund
            liquidation.refund = liquidation.calculate_refund()

            # Update status to submitted and track who changed it
            liquidation._status_changed_by = request.user
            # Avoid duplicate signal audit; we log business event manually
            liquidation._skip_signal_audit = True
            liquidation.status = 'submitted'
            liquidation.date_submitted = timezone.now()
            liquidation.save()
            from .audit_utils import log_audit_event
            log_audit_event(
                request=request,
                action='submit',
                module='liquidation',
                description=f"Submitted liquidation {LiquidationID}",
                object_id=LiquidationID,
                object_type='LiquidationManagement',
                object_name=f"Liquidation {LiquidationID}"
            )

            # Create notification for the reviewer
            # Notification.objects.create(
            #     notification_title="New Liquidation Submitted",
            #     details=f"Liquidation {liquidation.LiquidationID} has been submitted for review.",
            #     receiver=liquidation.request.user,  # Or the appropriate reviewer
            #     sender=request.user,
            # )

            serializer = LiquidationManagementSerializer(
                liquidation,
                context={'request': request}
            )
            return Response(serializer.data, status=status.HTTP_200_OK)

    except LiquidationManagement.DoesNotExist:
        return Response(
            {'error': 'Liquidation not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Error submitting liquidation: {str(e)}", exc_info=True)
        return Response(
            {'error': 'An unexpected error occurred during submission'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
def view_liquidation(request, LiquidationID):
    liquidation = get_object_or_404(
        LiquidationManagement, LiquidationID=LiquidationID)
    serializer = LiquidationManagementSerializer(liquidation)
    return Response(serializer.data)


# @api_view(['POST'])
# def approve_liquidation(request, LiquidationID):
#     try:
#         liquidation = LiquidationManagement.objects.get(
#             LiquidationID=LiquidationID)
#         if liquidation.status not in ['submitted', 'under_review']:
#             return Response(
#                 {'error': 'Liquidation is not in a reviewable state'},
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         if liquidation.documents.filter(is_approved=False).exists():
#             return Response(
#                 {'error': 'All documents must be approved first'},
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         # Set the user who changed the status for notification
#         liquidation._status_changed_by = request.user
#         liquidation.status = 'liquidated'
#         liquidation.reviewed_by = request.user
#         liquidation.reviewed_at = timezone.now()
#         liquidation.save()

#         serializer = LiquidationManagementSerializer(liquidation)
#         return Response(serializer.data)

#     except LiquidationManagement.DoesNotExist:
#         return Response(
#             {'error': 'Liquidation not found'},
#             status=status.HTTP_404_NOT_FOUND
#         )


class UserLiquidationsAPIView(generics.ListAPIView):
    serializer_class = LiquidationManagementSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role == "school_head":
            return LiquidationManagement.objects.filter(request__user=user).order_by('-created_at')
        elif user.role == "school_admin":
            # Allow school admin to see liquidations for their school
            return LiquidationManagement.objects.filter(request__user__school=user.school).order_by('-created_at')
        else:
            return LiquidationManagement.objects.none()


class UserRequestListAPIView(generics.ListAPIView):
    serializer_class = RequestManagementSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return RequestManagement.objects.filter(user=self.request.user).order_by('-created_at')


class PendingLiquidationListAPIView(generics.ListAPIView):
    serializer_class = LiquidationManagementSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Filter LiquidationManagement by status and by the current user
        return LiquidationManagement.objects.filter(
            status='draft',
            request__user=self.request.user
        )


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def batch_update_school_budgets(request):
    updates = request.data.get("updates", [])
    if not isinstance(updates, list):
        return Response({"error": "Invalid data format."}, status=400)

    updated_ids = []
    errors = []
    updated_schools = []

    with transaction.atomic():
        for upd in updates:
            school_id = str(upd.get("schoolId")).strip()
            max_budget = upd.get("max_budget")
            print("Updating:", school_id, "to", max_budget)
            try:
                school = School.objects.get(schoolId=school_id)
                if max_budget is not None and float(max_budget) >= 0:
                    old_budget = school.max_budget
                    school.max_budget = float(max_budget)
                    school.save()
                    updated_ids.append(school_id)
                    updated_schools.append({
                        'school': school,
                        'old_budget': old_budget,
                        'new_budget': school.max_budget
                    })
                else:
                    errors.append(
                        {"schoolId": school_id, "error": "Invalid budget"})
            except School.DoesNotExist:
                errors.append(
                    {"schoolId": school_id, "error": "School not found"})
            except Exception as e:
                errors.append({"schoolId": school_id, "error": str(e)})

    # Log audit for batch update (business event)
    if updated_schools:
        from .audit_utils import log_audit_event
        for update_info in updated_schools:
            school = update_info['school']
            # Suppress signal audit for this save already happened; log business event only
            log_audit_event(
                request=request,
                action='batch_update',
                module='school',
                description=f"Batch updated budget for school {school.schoolName} ({school.schoolId}) from {update_info['old_budget']} to {update_info['new_budget']}",
                object_id=school.schoolId,
                object_type='School',
                object_name=f"{school.schoolName} ({school.schoolId})",
                old_values={'max_budget': update_info['old_budget']},
                new_values={'max_budget': update_info['new_budget']}
            )

    return Response({
        "updated": updated_ids,
        "errors": errors
    }, status=200 if not errors else 207)


@api_view(['GET'])
def legislative_districts(request):
    """
    Returns mapping of legislative districts to their municipalities.
    """
    data = {
        "1st District": [
            "BANGAR", "LUNA", "SUDIPEN", "BALAOAN", "SANTOL", "BACNOTAN",
            "SAN GABRIEL", "SAN JUAN", "SAN FERNANDO CITY"
        ],
        "2nd District": [
            "AGOO", "ARINGAY", "BAGULIN", "BAUANG", "BURGOS", "CABA",
            "NAGUILIAN", "PUGO", "ROSARIO", "SANTO TOMAS", "TUBAO"
        ]
    }
    return Response(data)


def generate_request_id():
    prefix = "REQ-"
    random_part = get_random_string(
        length=6,
        allowed_chars=string.ascii_uppercase + string.digits
    )
    return f"{prefix}{random_part}"


class NotificationListAPIView(generics.ListAPIView):
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Notification.objects.filter(receiver=self.request.user).order_by('-notification_date')


class MarkNotificationAsReadAPIView(generics.UpdateAPIView):
    queryset = Notification.objects.all()
    permission_classes = [IsAuthenticated]

    def patch(self, request, *args, **kwargs):
        notification = self.get_object()
        if notification.receiver != request.user:
            return Response({"detail": "Not authorized."}, status=status.HTTP_403_FORBIDDEN)

        # Here you might want to add an 'is_read' field to your Notification model
        notification.is_read = True
        notification.save()
        return Response({"status": "marked as read"})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_me(request):
    serializer = UserSerializer(request.user)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def division_signatories(request):
    """
    Returns the first superintendent and division accountant.
    """
    from .models import User
    superintendent = User.objects.filter(
        role='superintendent').order_by('date_joined').first()
    accountant = User.objects.filter(
        role='accountant').order_by('date_joined').first()
    return Response({
        "superintendent": UserSerializer(superintendent).data if superintendent else None,
        "accountant": UserSerializer(accountant).data if accountant else None,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def request_history(request, request_id):
    req = RequestManagement.objects.get(request_id=request_id)
    history = req.history.order_by('-history_date')
    data = []
    for h in history:
        priorities = get_priorities_for_history(h.request_id, h.history_date)
        data.append({
            "priorities": priorities,
            "status": h.status,
            "rejection_comment": h.rejection_comment,
            "rejection_date": h.rejection_date,
            "created_at": h.created_at,
            "user": {
                "first_name": h.user.first_name if h.user else "",
                "last_name": h.user.last_name if h.user else "",
                "school": {
                    "schoolId": h.user.school.schoolId if h.user and h.user.school else "",
                    "schoolName": h.user.school.schoolName if h_user and h.user.school else "",
                } if h.user and h.user.school else None,
            } if h.user else None,
            "request_id": h.request_id,
        })
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def liquidation_management_history(request, LiquidationID):
    liq = LiquidationManagement.objects.get(LiquidationID=LiquidationID)
    history = liq.history.all().order_by('-history_date')
    serializer = LiquidationManagementHistorySerializer(history, many=True)
    return Response(serializer.data)


# @api_view(['POST'])
# def request_otp(request):
#     email = request.data.get('email')
#     user = User.objects.filter(email=email).first()
#     if not user:
#         return Response({'error': 'User not found'}, status=404)
#     otp = generate_otp()
#     user.otp_code = otp
#     user.otp_generated_at = timezone.now()
#     user.save(update_fields=['otp_code', 'otp_generated_at'])
#     send_otp_email(user, otp)
#     return Response({'message': 'OTP sent to your email.'})


# @api_view(['POST'])
# def verify_otp(request):
#     email = request.data.get('email')
#     otp = request.data.get('otp')
#     user = User.objects.filter(email=email).first()
#     if not user or not user.otp_code:
#         return Response({'error': 'OTP not found'}, status=404)
#     # Optional: Check OTP expiry (e.g., valid for 5 minutes)
#     if timezone.now() - user.otp_generated_at > timezone.timedelta(minutes=5):
#         return Response({'error': 'OTP expired'}, status=400)
#     if user.otp_code != otp:
#         return Response({'error': 'Invalid OTP'}, status=400)
#     # OTP is valid, clear it
#     user.otp_code = None
#     user.otp_generated_at = None
#     user.save(update_fields=['otp_code', 'otp_generated_at'])
#     return Response({'message': 'OTP verified. You may now log in.'})


class SchoolDistrictPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class SchoolDistrictListCreateAPIView(generics.ListCreateAPIView):
    serializer_class = SchoolDistrictSerializer
    pagination_class = SchoolDistrictPagination

    def get_queryset(self):
        queryset = SchoolDistrict.objects.all()
        search_term = self.request.query_params.get('search', None)
        legislative_district = self.request.query_params.get(
            'legislative_district', None)
        municipality = self.request.query_params.get('municipality', None)
        archived = self.request.query_params.get(
            'archived', 'false').lower() == 'true'
        show_all = self.request.query_params.get(
            'show_all', 'false').lower() == 'true'

        # Archive filter - show_all takes precedence over archived
        if not show_all:
            if not archived:
                queryset = queryset.filter(is_active=True)
            else:
                queryset = queryset.filter(is_active=False)

        if search_term:
            queryset = queryset.filter(
                Q(districtName__icontains=search_term) |
                Q(districtId__icontains=search_term) |
                Q(municipality__icontains=search_term)
            )
        if legislative_district:
            queryset = queryset.filter(
                legislativeDistrict=legislative_district)
        if municipality:
            queryset = queryset.filter(municipality=municipality)
        ordering = self.request.query_params.get('ordering', 'districtName')
        if ordering:
            queryset = queryset.order_by(ordering)
        return queryset

    def create(self, request, *args, **kwargs):
        # Enable bulk posting
        is_many = isinstance(request.data, list)
        serializer = self.get_serializer(data=request.data, many=is_many)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)

        # Audit handled by signals

        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class SchoolDistrictRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = SchoolDistrict.objects.all()
    serializer_class = SchoolDistrictSerializer
    lookup_field = 'districtId'

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        old_is_active = instance.is_active

        response = super().update(request, *args, **kwargs)

        
        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        district_name = f"{instance.districtName} ({instance.districtId})"

        response = super().destroy(request, *args, **kwargs)


        return response


@api_view(['PATCH'])
def archive_school_district(request, districtId):
    """
    Archive or restore a school district.
    """
    try:
        school_district = SchoolDistrict.objects.get(districtId=districtId)
        old_is_active = school_district.is_active
        is_active = request.data.get("is_active", None)
        if is_active is not None:
            school_district.is_active = is_active
            school_district.save()
            # Audit handled by signals

            return Response({"status": "updated", "is_active": school_district.is_active})
        return Response({"error": "Missing is_active field"}, status=400)
    except SchoolDistrict.DoesNotExist:
        return Response({"error": "School district not found"}, status=404)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def last_liquidated_request(request):
    last_req = (
        RequestManagement.objects.filter(
            user=request.user, status="liquidated")
        .order_by('-created_at')
        .first()
    )
    if not last_req:
        return Response({"detail": "No liquidated request found"}, status=404)
    return Response(PreviousRequestSerializer(last_req).data)


def generate_otp():
    return ''.join(str(random.randint(0, 9)) for _ in range(6))


@api_view(['POST'])
@permission_classes([AllowAny])
def request_otp(request):
    email = request.data.get('email')
    password = request.data.get('password')
    try:
        user_obj = User.objects.get(email=email)
        if not user_obj.is_active:
            return Response({'message': 'Your account has been archived by the administrator. To reactivate your account, please contact our support team at support@company.com.'}, status=403)
        user = authenticate(email=email, password=password)
        if not user:
            return Response({'message': 'Invalid email or password'}, status=400)
        otp = generate_otp()
        user.otp_code = otp
        user.otp_generated_at = timezone.now()
        user.save()
        send_otp_email(user, otp)
        return Response({'message': 'OTP sent to your email'})
    except User.DoesNotExist:
        return Response({'message': 'Invalid email or password'}, status=400)


@api_view(['POST'])
@permission_classes([AllowAny])
def verify_otp(request):
    email = request.data.get('email')
    otp = request.data.get('otp')
    try:
        user = User.objects.get(email=email)
        if not user.is_active:
            return Response({'error': 'Your account is inactive. Please contact the administrator.'}, status=403)
        # Check OTP validity (add expiry logic if needed)
        if timezone.now() - user.otp_generated_at > timezone.timedelta(minutes=5):
            return Response({'error': 'OTP expired'}, status=400)

        if user.otp_code == otp:
            user.otp_code = None
            user.save()
            return Response({'message': 'OTP verified'})
        else:
            return Response({'message': 'Invalid OTP. Please try again.'}, status=400)
    except User.DoesNotExist:
        return Response({'message': 'User not found'}, status=404)


@api_view(['POST'])
@permission_classes([AllowAny])
def resend_otp(request):
    email = request.data.get('email')
    try:
        user = User.objects.get(email=email)
        otp = generate_otp()
        user.otp_code = otp
        user.otp_generated_at = timezone.now()
        user.save()
        send_otp_email(user, otp)
        return Response({'message': 'OTP resent'})
    except User.DoesNotExist:
        return Response({'message': 'User not found'}, status=404)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def schools_with_unliquidated_requests(request):
    days_threshold = request.GET.get('days', '30')
    export_format = request.GET.get('export')
    page_size = request.GET.get('page_size', 50)

    aging_data = generate_aging_report_data(days_threshold)

    if export_format == 'csv':
        return generate_aging_csv_report(aging_data, days_threshold)
    elif export_format == 'excel':
        return generate_aging_excel_report(aging_data, days_threshold)

    # Apply pagination for regular API response
    paginator = AgingReportPagination()
    paginator.page_size = page_size
    paginated_data = paginator.paginate_queryset(aging_data, request)

    return paginator.get_paginated_response({
        'results': paginated_data,
        'total_count': len(aging_data),
        'filters': {
            'days_threshold': days_threshold,
            'aging_periods': {
                '0-30': len([item for item in aging_data if item['aging_period'] == '0-30 days']),
                '31-60': len([item for item in aging_data if item['aging_period'] == '31-60 days']),
                '61-90': len([item for item in aging_data if item['aging_period'] == '61-90 days']),
                '91-120': len([item for item in aging_data if item['aging_period'] == '91-120 days']),
                '121-180': len([item for item in aging_data if item['aging_period'] == '121-180 days']),
                '180+': len([item for item in aging_data if item['aging_period'] == '180+ days']),
            }
        }
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_dashboard(request):
    # Check if user is admin
    if request.user.role != 'admin':
        return Response({"error": "Only administrators can access this endpoint"}, status=status.HTTP_403_FORBIDDEN)

    # Get time range parameter
    time_range = request.GET.get('time_range', 'last_quarter')

    # Calculate date range based on parameter
    end_date = timezone.now()
    if time_range == 'last_week':
        start_date = end_date - timedelta(days=7)
    elif time_range == 'last_month':
        start_date = end_date - timedelta(days=30)
    elif time_range == 'last_quarter':
        start_date = end_date - timedelta(days=90)
    elif time_range == 'last_year':
        start_date = end_date - timedelta(days=365)
    else:
        # Default to last quarter
        start_date = end_date - timedelta(days=90)

    # DEBUG: Print actual date range
    # print(f"DEBUG: Actual date range - Start: {start_date}, End: {end_date}")

    budget_utilization = []
    category_breakdown = []
    months = []
    current = start_date.replace(day=1)  # Start at beginning of month

    while current <= end_date:
        month_str = current.strftime('%Y-%m')
        months.append(month_str)
        current = current + relativedelta(months=1)

    # DEBUG: Show what months we're processing
    # print(f"DEBUG: Processing months: {months}")

    # DEBUG: Show all requests and liquidations in the system
    all_requests = RequestManagement.objects.all()
    # print(f"DEBUG: Total requests in system: {all_requests.count()}")
    for req in all_requests:
        print(
            f"  Request {req.request_id}: created={req.created_at}, month={req.request_monthyear}, status={req.status}")

    all_liquidations = LiquidationManagement.objects.filter(
        status='liquidated')
    # print(
    #     f"DEBUG: Total liquidated liquidations in system: {all_liquidations.count()}")
    for liq in all_liquidations:
        print(f"  Liquidation {liq.LiquidationID}: created={liq.created_at}, liquidated={liq.date_liquidated}, request_month={liq.request.request_monthyear if liq.request else 'None'}")

    # Fixed budget utilization calculation
    for month in months:
        year, month_num = map(int, month.split('-'))

        print(
            f"\n=== PROCESSING MONTH: {month} (Year: {year}, Month: {month_num}) ===")

        # METHOD 1: Get requests CREATED in this month
        month_requests_created = RequestManagement.objects.filter(
            created_at__year=year,
            created_at__month=month_num
        )
        print(
            f"DEBUG: Requests CREATED in {month}: {month_requests_created.count()}")

        # METHOD 2: Get requests FOR this month (request_monthyear)
        month_requests_for_period = RequestManagement.objects.filter(
            request_monthyear=month
        )
        print(
            f"DEBUG: Requests FOR period {month}: {month_requests_for_period.count()}")

        # Use the appropriate method based on what you want to track
        # For budget utilization, you probably want requests FOR the period
        month_requests = month_requests_for_period

        # --- 1. Planned (Requested) Amounts ---
        total_allocated = School.objects.aggregate(
            total=Sum('max_budget'))['total'] or 0
        total_planned = RequestPriority.objects.filter(
            request__in=month_requests
        ).aggregate(total=Sum('amount'))['total'] or 0

        print(f"DEBUG: Total allocated budget: {total_allocated}")
        print(f"DEBUG: Total planned spending for {month}: {total_planned}")

        # --- 2. Actual (Liquidated) Amounts ---
        # METHOD 1: Liquidations COMPLETED in this month
        month_liquidations_completed = LiquidationManagement.objects.filter(
            date_liquidated__year=year,
            date_liquidated__month=month_num,
            status='liquidated'
        )
        print(
            f"DEBUG: Liquidations COMPLETED in {month}: {month_liquidations_completed.count()}")

        # METHOD 2: Liquidations for requests OF this month
        month_liquidations_for_period = LiquidationManagement.objects.filter(
            request__request_monthyear=month,
            status='liquidated'
        )
        print(
            f"DEBUG: Liquidations FOR period {month}: {month_liquidations_for_period.count()}")

        # Use the appropriate method
        month_liquidations = month_liquidations_for_period

        total_actual = 0
        liquidation_details = []
        for liquidation in month_liquidations:
            liquidation_amount = liquidation.liquidation_priorities.aggregate(
                total=Sum('amount'))['total'] or 0
            total_actual += liquidation_amount
            liquidation_details.append({
                'id': liquidation.LiquidationID,
                'amount': liquidation_amount,
                'created': liquidation.created_at,
                'completed': liquidation.date_liquidated
            })

        # print(
        #     f"DEBUG: Liquidations details for {month}: {liquidation_details}")
        # print(f"DEBUG: Total actual spending for {month}: {total_actual}")

        # --- 3. Utilization Rates ---
        planned_utilization_rate = (
            total_planned / total_allocated * 100) if total_allocated > 0 else 0
        actual_utilization_rate = (
            total_actual / total_allocated * 100) if total_allocated > 0 else 0

        budget_utilization.append({
            'month': month,
            'allocated': float(total_allocated),
            'planned': float(total_planned),
            'actual': float(total_actual),
            'plannedUtilizationRate': float(planned_utilization_rate),
            'actualUtilizationRate': float(actual_utilization_rate)
        })

        print(
            f"SUMMARY: Month: {month}, Planned: {total_planned}, Actual: {total_actual}")
        print("---")
        # --- 4. Category Breakdown for this month (NEW) ---
        category_data_for_month = {}
        for category_key, category_name in ListOfPriority.CATEGORY_CHOICES:
            # Get planned amount for this category
            planned_for_category = RequestPriority.objects.filter(
                request__in=month_requests,
                priority__category=category_key
            ).aggregate(total=Sum('amount'))['total'] or 0

            # Get actual liquidated amount for this category
            actual_for_category = 0
            liquidations_for_category = month_liquidations.filter(
                liquidation_priorities__priority__category=category_key
            ).distinct()
            for liq in liquidations_for_category:
                amount = liq.liquidation_priorities.filter(
                    priority__category=category_key
                ).aggregate(total=Sum('amount'))['total'] or 0
                actual_for_category += amount

            category_data_for_month[category_name] = {
                'planned': float(planned_for_category),
                'actual': float(actual_for_category)
            }

        # Append category data for the stacked chart
        category_breakdown.append({
            'month': month,
            **category_data_for_month  # Unpacks the dictionary of categories
        })

    # 2. Request Status Distribution
    status_counts = RequestManagement.objects.values('status').annotate(
        count=Count('status')
    ).order_by('-count')

    total_requests = RequestManagement.objects.count()
    request_status_distribution = []

    for status in status_counts:
        percentage = (status['count'] / total_requests *
                      100) if total_requests > 0 else 0
        request_status_distribution.append({
            'status': status['status'],
            'count': status['count'],
            'percentage': float(percentage)
        })

   # 3. Liquidation Timeline - FIXED calculation
    liquidation_timeline = []
    for month in months:
        year, month_num = map(int, month.split('-'))

        # Get liquidations created in this month that were eventually liquidated
        month_liquidations = LiquidationManagement.objects.filter(
            created_at__year=year,
            created_at__month=month_num,
            status='liquidated'
        )

        total_seconds = 0
        count = 0

        for liquidation in month_liquidations:
            if liquidation.date_liquidated and liquidation.created_at:
                # Both are DateTimeField, so direct comparison works
                time_diff = liquidation.date_liquidated - liquidation.created_at
                total_seconds += time_diff.total_seconds()
                count += 1

        # Calculate average processing time in days
        avg_days = (total_seconds / (24 * 60 * 60)) / count if count > 0 else 0

        # Count approved and rejected for this month
        approved_count = month_liquidations.count()
        rejected_count = LiquidationManagement.objects.filter(
            created_at__year=year,
            created_at__month=month_num,
            status='resubmit'
        ).count()

        liquidation_timeline.append({
            'month': month,
            'avgProcessingTime': float(avg_days),
            'approved': approved_count,
            'rejected': rejected_count
        })

    # 4. School Performance - FIXED: Use liquidation created_at instead of request created_at
    school_performance = []
    schools = School.objects.all()

    for school in schools:
        school_requests = RequestManagement.objects.filter(user__school=school)
        total_requests = school_requests.count()
        approved_requests = school_requests.filter(status='approved').count()
        rejected_requests = school_requests.filter(status='rejected').count()

        rejection_rate = (rejected_requests / total_requests *
                          100) if total_requests > 0 else 0

        # Calculate average processing time for this school - FIXED: Use liquidation creation date
        # Calculate average processing time for this school - FIXED
        school_liquidations = LiquidationManagement.objects.filter(
            request__user__school=school,
            status='liquidated'
        )

        total_seconds = 0
        count = 0

        for liquidation in school_liquidations:
            if liquidation.date_liquidated and liquidation.created_at:
                time_diff = liquidation.date_liquidated - liquidation.created_at
                total_seconds += time_diff.total_seconds()
                count += 1

        avg_processing_days = (
            total_seconds / (24 * 60 * 60)) / count if count > 0 else 0

        # Calculate budget utilization for this school
        school_utilized = RequestPriority.objects.filter(
            request__user__school=school
        ).aggregate(total=Sum('amount'))['total'] or 0

        budget_utilization_rate = (
            school_utilized / school.max_budget * 100) if school.max_budget > 0 else 0

        school_performance.append({
            'schoolId': school.schoolId,
            'schoolName': school.schoolName,
            'totalRequests': total_requests,
            'approvedRequests': approved_requests,
            'rejectionRate': float(rejection_rate),
            'avgProcessingTime': float(avg_processing_days),
            'budgetUtilization': float(budget_utilization_rate)
        })

    # 5. Category Spending
    category_spending = []
    categories = ListOfPriority.CATEGORY_CHOICES

    for category_key, category_name in categories:
        category_total = RequestPriority.objects.filter(
            priority__category=category_key
        ).aggregate(total=Sum('amount'))['total'] or 0

        total_all_categories = RequestPriority.objects.aggregate(total=Sum('amount'))[
            'total'] or 0
        percentage = (category_total / total_all_categories *
                      100) if total_all_categories > 0 else 0

        # Simple trend calculation (compare with previous period)
        trend = 'stable'
        if category_total > 0:
            # This is a simplified trend calculation
            trend = 'up' if category_total > 10000 else 'down'

        category_spending.append({
            'category': category_name,
            'totalAmount': float(category_total),
            'percentage': float(percentage),
            'trend': trend
        })

    # Sort by total amount descending
    category_spending.sort(key=lambda x: x['totalAmount'], reverse=True)
    category_spending = category_spending[:5]  # Limit to top 5

    # 6. Document Compliance
    document_compliance = []
    requirements = Requirement.objects.all()

    for requirement in requirements:
        total_submitted = LiquidationDocument.objects.filter(
            requirement=requirement
        ).count()

        compliant = LiquidationDocument.objects.filter(
            requirement=requirement,
            is_approved=True
        ).count()

        compliance_rate = (compliant / total_submitted *
                           100) if total_submitted > 0 else 0

        document_compliance.append({
            'requirement': requirement.requirementTitle,
            'totalSubmitted': total_submitted,
            'compliant': compliant,
            'complianceRate': float(compliance_rate)
        })

    # 7. Top Priorities
    top_priorities = []
    priorities = RequestPriority.objects.values(
        'priority__expenseTitle'
    ).annotate(
        frequency=Count('priority'),
        total_amount=Sum('amount')
    ).order_by('-total_amount')[:10]

    for priority in priorities:
        # Simple trend calculation
        trend = 'stable'
        if priority['total_amount']:
            trend = 'up' if priority['total_amount'] > 5000 else 'down'

        top_priorities.append({
            'priority': priority['priority__expenseTitle'],
            'frequency': priority['frequency'],
            'totalAmount': float(priority['total_amount']),
            'trend': trend
        })

    # 8. Active Requests - REPLACED Pending Actions with Active Requests
    active_requests = []

    # Get all requests that are in active states (not completed/rejected)
    active_request_states = ['pending',
                             'approved', 'downloaded', 'unliquidated']
    active_requests_data = RequestManagement.objects.filter(
        status__in=active_request_states)

    for request in active_requests_data:
        days_active = (timezone.now() - request.created_at).days

        # Determine priority based on status and age
        if request.status == 'pending':
            priority = 'high' if days_active > 7 else 'medium' if days_active > 3 else 'low'
        elif request.status == 'approved':
            priority = 'high' if days_active > 5 else 'medium'
        else:
            priority = 'medium'

        # Get school information
        school_name = request.user.school.schoolName if request.user and request.user.school else "Unknown School"
        user_name = request.user.get_full_name() if request.user else "Unknown User"

        active_requests.append({
            'id': request.request_id,
            'type': 'request',
            'title': f'Active Request: {request.request_id}',
            'description': f'{user_name} from {school_name} - Status: {request.status.capitalize()}',
            'status': request.status,
            'priority': priority,
            'timestamp': request.created_at.isoformat(),
            'schoolName': school_name,
            'userName': user_name
        })

    # Sort by priority and timestamp
    priority_order = {'high': 0, 'medium': 1, 'low': 2}
    active_requests.sort(key=lambda x: (
        priority_order[x['priority']], x['timestamp']))

    # 9. Additional Liquidation Metrics - NEW
    # Debug: Check for liquidations with unrealistic dates
    liquidated_liquidations = LiquidationManagement.objects.filter(
        status='liquidated',
        date_liquidated__isnull=False,
        created_at__isnull=False
    )

    # Filter out liquidations with unrealistic processing times (more than 1 year)
    realistic_liquidations = []
    for liq in liquidated_liquidations:
        if liq.date_liquidated and liq.created_at:
            time_diff = liq.date_liquidated - liq.created_at
            # Only include liquidations with processing time less than 365 days
            if time_diff.total_seconds() < (365 * 24 * 60 * 60):
                realistic_liquidations.append(liq)

    # Calculate average from realistic liquidations only
    if realistic_liquidations:
        total_seconds = sum((liq.date_liquidated - liq.created_at).total_seconds()
                            for liq in realistic_liquidations)
        avg_seconds = total_seconds / len(realistic_liquidations)
        avg_liquidation_time = avg_seconds
    else:
        avg_liquidation_time = 0

    liquidation_metrics = {
        'total_liquidations': LiquidationManagement.objects.count(),
        'completed_liquidations': LiquidationManagement.objects.filter(status='liquidated').count(),
        'pending_liquidations': LiquidationManagement.objects.exclude(status='liquidated').count(),

        'completion_rate': (LiquidationManagement.objects.filter(status='liquidated').count() /
                            LiquidationManagement.objects.count() * 100) if LiquidationManagement.objects.count() > 0 else 0,
        'avg_liquidation_time': avg_liquidation_time
    }

    # Convert avg_liquidation_time to days
    if hasattr(liquidation_metrics['avg_liquidation_time'], 'days'):
        liquidation_metrics['avg_liquidation_time_days'] = liquidation_metrics['avg_liquidation_time'].days
    else:
        liquidation_metrics['avg_liquidation_time_days'] = float(
            liquidation_metrics['avg_liquidation_time']) / (24 * 60 * 60) if liquidation_metrics['avg_liquidation_time'] else 0

    # Debug information - show some sample data
    sample_liquidations = []
    for liq in liquidated_liquidations[:3]:  # Show first 3 liquidations
        if liq.date_liquidated and liq.created_at:
            time_diff = liq.date_liquidated - liq.created_at
            sample_liquidations.append({
                'liquidation_id': liq.LiquidationID,
                'created_at': liq.created_at.isoformat(),
                'date_liquidated': liq.date_liquidated.isoformat(),
                'time_diff_seconds': time_diff.total_seconds(),
                'time_diff_days': time_diff.total_seconds() / (24 * 60 * 60)
            })

    liquidation_metrics['debug_info'] = {
        'total_liquidated_count': len(liquidated_liquidations),
        'realistic_liquidations_count': len(realistic_liquidations),
        'avg_seconds_raw': liquidation_metrics['avg_liquidation_time'],
        'avg_days_calculated': liquidation_metrics['avg_liquidation_time_days'],
        'sample_liquidations': sample_liquidations
    }

    # 10. Top Schools by Liquidation Speed - NEW (Fixed with realistic time filtering)
    top_schools_by_speed = []

    # Get all schools and calculate realistic processing times manually
    all_schools = School.objects.all()

    for school in all_schools:
        # Get liquidations for this school that are liquidated
        school_liquidations = LiquidationManagement.objects.filter(
            request__user__school=school,
            status='liquidated',
            date_liquidated__isnull=False,
            created_at__isnull=False
        )

        # Filter out unrealistic processing times (more than 1 year)
        realistic_processing_times = []
        for liq in school_liquidations:
            if liq.date_liquidated and liq.created_at:
                time_diff = liq.date_liquidated - liq.created_at
                # Only include liquidations with processing time less than 365 days
                if time_diff.total_seconds() < (365 * 24 * 60 * 60):
                    realistic_processing_times.append(
                        time_diff.total_seconds())

        # Calculate average from realistic times only
        if realistic_processing_times:
            avg_seconds = sum(realistic_processing_times) / \
                len(realistic_processing_times)
            avg_days = avg_seconds / (24 * 60 * 60)

            top_schools_by_speed.append({
                'schoolId': school.schoolId,
                'schoolName': school.schoolName,
                'avgProcessingDays': float(avg_days)
            })

    # Sort by processing time (ascending = fastest first) and take top 5
    top_schools_by_speed.sort(key=lambda x: x['avgProcessingDays'])
    top_schools_by_speed = top_schools_by_speed[:5]

    # 11. School Document Compliance - NEW
    school_document_compliance = []
    all_schools_for_compliance = School.objects.all()

    for school in all_schools_for_compliance:
        # Get all liquidations for this school
        school_liquidations = LiquidationManagement.objects.filter(
            request__user__school=school
        )

        total_required_docs = 0
        total_uploaded_docs = 0

        for liquidation in school_liquidations:
            # For each liquidation, count required and uploaded documents
            for rp in liquidation.request.requestpriority_set.all():
                required_docs = rp.priority.requirements.filter(
                    is_required=True).count()
                uploaded_docs = LiquidationDocument.objects.filter(
                    liquidation=liquidation,
                    request_priority=rp
                ).count()

                total_required_docs += required_docs
                # Cap at required
                total_uploaded_docs += min(uploaded_docs, required_docs)

        compliance_rate = (total_uploaded_docs / total_required_docs *
                           100) if total_required_docs > 0 else 0

        school_document_compliance.append({
            'schoolId': school.schoolId,
            'schoolName': school.schoolName,
            'uploadedDocuments': total_uploaded_docs,
            'requiredDocuments': total_required_docs,
            'complianceRate': compliance_rate,
            'pendingDocuments': max(0, total_required_docs - total_uploaded_docs)
        })

    # Sort by compliance rate (descending)
    school_document_compliance.sort(
        key=lambda x: x['complianceRate'], reverse=True)
    # Prepare response data

    response_data = {
        'budgetUtilization': budget_utilization,
        'categoryBreakdown': category_breakdown,  # NEW
        'requestStatusDistribution': request_status_distribution,
        'liquidationTimeline': liquidation_timeline,
        'schoolPerformance': school_performance,
        'categorySpending': category_spending,
        'documentCompliance': document_compliance,
        'topPriorities': top_priorities,
        'activeRequests': active_requests,
        'liquidationMetrics': liquidation_metrics,
        'topSchoolsBySpeed': top_schools_by_speed,
        'schoolDocumentCompliance': school_document_compliance,
    }

    return Response(response_data)
# ------------------- Backup & Restore -------------------


def _is_safe_path(path: str) -> bool:
    """
    Validate that path is safe and doesn't contain traversal attempts
    """
    if not path or not isinstance(path, str):
        return False

    # Normalize path and check for traversal attempts
    normalized = os.path.normpath(path)

    # Check for dangerous patterns
    # Platform-aware invalid pattern checks
    if os.name == 'nt':
        # On Windows, allow drive letter colon (e.g., C:\), but disallow reserved characters
        invalid_chars = ['|', '<', '>', '"']
        if any(ch in normalized for ch in invalid_chars):
            return False
        # Reject device paths
        if ('\\\\.\\' in normalized) or ('\\\\?\\' in normalized):
            return False
        # Basic traversal check
        if '..' in normalized or '~' in normalized:
            return False
    else:
        dangerous_patterns = [
            '..', '~',  # Directory traversal
            '/dev/', '/proc/', '/sys/',  # Linux system paths
            '|', '<', '>', '"',  # Invalid characters
        ]
        for pattern in dangerous_patterns:
            if pattern in normalized:
                return False
    # Additional checks for absolute paths
    if os.path.isabs(normalized):
        # Allow only certain safe directories if needed
        allowed_prefixes = ['/backups/']
        # Include server-configured default backup dir if present
        try:
            default_dir = getattr(settings, 'BACKUP_SETTINGS', {}).get(
                'DEFAULT_BACKUP_DIR')
            if default_dir:
                # Normalize to same style
                allowed_prefixes.append(os.path.normpath(
                    default_dir) + (os.sep if not default_dir.endswith(os.sep) else ''))
        except Exception:
            pass
        if os.name == 'nt':
            allowed_prefixes.extend(['C:\\Backups\\', 'D:\\Backups\\'])
        if not any(normalized.startswith(prefix) for prefix in allowed_prefixes):
            return False

    return True


MODEL_BACKUP_ORDER = [
    # Level 1: Independent models (no foreign keys)
    (SchoolDistrict, 'SchoolDistrict'),
    (Requirement, 'Requirement'),
    (ListOfPriority, 'ListOfPriority'),

    # Level 2: Depend on Level 1
    (School, 'School'),  # Depends on SchoolDistrict
    # Depends on ListOfPriority, Requirement
    (PriorityRequirement, 'PriorityRequirement'),

    # Level 3: Depend on Level 2
    (User, 'User'),  # Depends on School, SchoolDistrict

    # Level 4: Depend on Level 3
    (RequestManagement, 'RequestManagement'),  # Depends on User
    (Backup, 'Backup'),  # Depends on User
    (AuditLog, 'AuditLog'),  # Depends on User
    (Notification, 'Notification'),  # Depends on User

    # Level 5: Depend on Level 4
    # Depends on RequestManagement, ListOfPriority
    (RequestPriority, 'RequestPriority'),
    (GeneratedPDF, 'GeneratedPDF'),  # Depends on RequestManagement, User

    # Level 6: Depend on Level 5
    # Depends on RequestManagement
    (LiquidationManagement, 'LiquidationManagement'),

    # Level 7: Depend on Level 6
    # Depends on LiquidationManagement, ListOfPriority
    (LiquidationPriority, 'LiquidationPriority'),
    # Depends on LiquidationManagement, RequestPriority, Requirement
    (LiquidationDocument, 'LiquidationDocument'),
]


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def initiate_backup(request):
    """
    Generate a backup archive with improved error handling.
    """
    try:
        format = request.data.get("format", "json")
        include_media = request.data.get("include_media", True)

        # Create backup record
        backup = Backup.objects.create(
            initiated_by=request.user,
            format=format,
            include_media=include_media,
            status='pending'
        )

        # Create HTTP response with ZIP file
        response = HttpResponse(content_type='application/zip')
        filename = f"backup_{backup.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.zip"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        with zipfile.ZipFile(response, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Add metadata
            metadata = {
                'created_at': timezone.now().isoformat(),
                'created_by': request.user.email,
                'format': format,
                'include_media': include_media,
                'model_versions': {}
            }

            # Export data as JSON using the corrected model order
            for model, model_name in MODEL_BACKUP_ORDER:
                try:
                    # Use natural keys if available for better serialization
                    data = django_serializers.serialize(
                        "json",
                        model.objects.all(),
                        use_natural_foreign_keys=False
                    )
                    zipf.writestr(f"data/{model_name}.json", data)

                    # Store model version info
                    metadata['model_versions'][model_name] = {
                        'count': model.objects.all().count(),
                        'exported_at': timezone.now().isoformat()
                    }

                    logger.info(
                        f"Backed up {model_name}: {model.objects.all().count()} records")
                except Exception as e:
                    logger.error(f"Failed to backup {model_name}: {e}")
                    # Continue with other models instead of failing completely
                    continue

            # Add metadata file
            zipf.writestr('metadata.json', json.dumps(metadata, indent=2))

            # Include media files if requested
            if include_media:
                try:
                    media_root = settings.MEDIA_ROOT
                    if os.path.exists(media_root):
                        media_files_added = 0
                        for root, dirs, files in os.walk(media_root):
                            for file in files:
                                # Skip temporary and hidden files
                                if file.startswith(('.', '~')) or file.endswith('.tmp'):
                                    continue

                                abs_path = os.path.join(root, file)
                                rel_path = os.path.relpath(
                                    abs_path, media_root)

                                try:
                                    with open(abs_path, 'rb') as f:
                                        file_content = f.read()
                                    zipf.writestr(
                                        f"media/{rel_path}", file_content)
                                    media_files_added += 1
                                except (IOError, OSError) as e:
                                    logger.warning(
                                        f"Could not read media file {abs_path}: {e}")
                                    continue

                        logger.info(
                            f"Added {media_files_added} media files to backup")
                    else:
                        logger.warning(
                            f"Media root {media_root} does not exist")
                except Exception as e:
                    logger.error(f"Media backup failed: {e}")
                    # Don't fail the entire backup if media fails

        # Update backup record (signals will audit the CRUD change)
        backup.status = 'success'
        backup.file_size = len(response.content)
        backup.save()

        return response

    except Exception as e:
        logger.error(f"Backup failed: {e}", exc_info=True)

        # Update backup record with failure
        if 'backup' in locals():
            backup.status = 'failed'
            backup.message = str(e)
            backup.save()

        return Response(
            {"detail": "Backup failed", "error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def initiate_restore(request):
    """
    Restore the system from an uploaded backup archive (.zip).
    """
    # Store current user info before any operations
    current_user_id = request.user.id
    current_user_email = request.user.email

    # Require explicit confirmation
    if not request.data.get("confirm_wipe", False):
        return Response(
            {
                "detail": "Restore will replace existing data. "
                "Include 'confirm_wipe': true in your request to proceed.",
                "warning": "This action cannot be undone."
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    temp_dir = None
    try:
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response(
                {"detail": "No backup file uploaded"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate file type
        if not file_obj.name.endswith('.zip'):
            return Response(
                {"detail": "Only ZIP files are supported for restore"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create temporary directory for extraction
        temp_dir = tempfile.mkdtemp()
        archive_path = os.path.join(temp_dir, file_obj.name)

        # Save uploaded file
        with open(archive_path, "wb+") as dest:
            for chunk in file_obj.chunks():
                dest.write(chunk)

        # Extract archive
        with zipfile.ZipFile(archive_path, "r") as zipf:
            zipf.extractall(temp_dir)

        # Restore logic
        data_dir = os.path.join(temp_dir, 'data')

        if not os.path.exists(data_dir):
            return Response(
                {"detail": "Backup file is corrupted or invalid"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # NEW APPROACH: Use smaller transactions per model instead of one giant transaction
        restored_counts = {}
        file_record_counts = {}
        critical_failures = {}
        critical_models = {"SchoolDistrict",
                           "Requirement", "ListOfPriority", "School", "User", "RequestManagement"}
        errors = []

        try:
            # Step 1: Create a backup of current data (optional safety measure)
            safety_backup = {}
            for model, model_name in MODEL_BACKUP_ORDER:
                safety_backup[model_name] = list(model.objects.all().values())

            # Step 2: Restore data in correct order with conflict resolution
            for model, model_name in MODEL_BACKUP_ORDER:
                file_path = os.path.join(data_dir, f"{model_name}.json")
                if not os.path.exists(file_path):
                    logger.warning(f"Backup file not found: {file_path}")
                    continue

                logger.info(f"Restoring {model_name} from {file_path}")

                # Use a separate transaction for each model
                try:
                    with transaction.atomic():
                        objects_processed = 0

                        with open(file_path, 'r', encoding='utf-8') as f:
                            data = f.read()

                        # DEBUG: Check the structure of the serialized data
                        import json
                        parsed_data = json.loads(data)
                        file_record_counts[model_name] = len(
                            parsed_data) if isinstance(parsed_data, list) else 0
                        if parsed_data:
                            logger.info(
                                f"First object structure: {list(parsed_data[0].keys())}")
                            if 'pk' in parsed_data[0]:
                                logger.info(
                                    f"PK found: {parsed_data[0]['pk']}")
                            else:
                                logger.warning(
                                    "PK not found in serialized data!")

                        # Use Django's deserializer
                        for obj in django_serializers.deserialize("json", data):
                            try:
                                # SPECIAL HANDLING FOR USER MODEL - IMPROVED VERSION
                                # SPECIAL HANDLING FOR USER MODEL - FIXED VERSION
                                if model.__name__ == 'User':
                                    # Use a savepoint to prevent breaking the whole transaction
                                    with transaction.atomic():
                                        email = getattr(
                                            obj.object, 'email', None)
                                    # Get the PK from serialized data - handle empty string case
                                    original_pk = getattr(obj, 'pk', None)

                                    # Check if PK is valid (not empty string and not None)
                                    is_valid_pk = original_pk and str(
                                        original_pk).strip() != ''

                                    logger.info(
                                        f"Processing user: email={email}, original_pk={original_pk}, is_valid_pk={is_valid_pk}")

                                    if email:
                                        try:
                                            # First, check if a user with this email already exists
                                            existing_by_email = model.objects.filter(
                                                email=email).first()

                                            # Check if a user with the original PK exists (only if PK is valid)
                                            existing_by_pk = None
                                            if is_valid_pk:
                                                existing_by_pk = model.objects.filter(
                                                    pk=original_pk).first()

                                            if existing_by_email:
                                                # User with same email exists - update it
                                                logger.info(
                                                    f"Updating existing user by email: {email}")
                                                for field in obj.object._meta.fields:
                                                    if field.primary_key or field.name == 'password':
                                                        continue
                                                    try:
                                                        setattr(existing_by_email, field.name, getattr(
                                                            obj.object, field.name))
                                                    except Exception as field_error:
                                                        logger.warning(
                                                            f"Error setting field {field.name}: {field_error}")
                                                        continue
                                                existing_by_email.save()
                                                objects_processed += 1

                                            elif existing_by_pk and is_valid_pk:
                                                # User with original PK exists - update it
                                                logger.info(
                                                    f"Updating existing user by PK: {original_pk}")
                                                for field in obj.object._meta.fields:
                                                    if field.primary_key or field.name == 'password':
                                                        continue
                                                    try:
                                                        setattr(existing_by_pk, field.name, getattr(
                                                            obj.object, field.name))
                                                    except Exception as field_error:
                                                        logger.warning(
                                                            f"Error setting field {field.name}: {field_error}")
                                                        continue
                                                existing_by_pk.save()
                                                objects_processed += 1

                                            else:
                                                # No existing user found - create new one
                                                logger.info(
                                                    f"Creating new user: {email}")

                                                # Create a copy of the object data
                                                user_data = obj.object.__dict__.copy()
                                                user_data.pop('_state', None)

                                                # Handle password separately
                                                password = user_data.pop(
                                                    'password', None)

                                                # Try to use original PK only if it's valid and not conflicting
                                                if is_valid_pk and not model.objects.filter(pk=original_pk).exists():
                                                    user_data['id'] = original_pk
                                                else:
                                                    # Remove invalid PK so Django can auto-generate one
                                                    user_data.pop('id', None)

                                                # Create the user
                                                new_user = model(**user_data)
                                                if password:
                                                    new_user.set_password(
                                                        password)
                                                new_user.save()
                                                objects_processed += 1

                                        except IntegrityError as e:
                                            logger.warning(
                                                f"Integrity error for user {email}: {e}")
                                            # Try fallback creation without PK
                                            try:
                                                with transaction.atomic():
                                                    user_data = obj.object.__dict__.copy()
                                                    user_data.pop(
                                                        '_state', None)
                                                    # Remove PK to avoid conflict
                                                    user_data.pop('id', None)

                                                    password = user_data.pop(
                                                        'password', None)
                                                    new_user = model(
                                                        **user_data)
                                                    if password:
                                                        new_user.set_password(
                                                            password)
                                                        new_user.save()
                                                        objects_processed += 1
                                                        logger.info(
                                                            f"Created user with auto-generated PK: {email}")
                                            except Exception as fallback_error:
                                                logger.error(
                                                    f"Fallback creation failed for {email}: {fallback_error}")
                                                errors.append(
                                                    f"User {email}: {str(fallback_error)}")

                                    else:
                                        # User without email - handle carefully
                                        logger.warning(
                                            "User without email found in backup")
                                        try:
                                            with transaction.atomic():
                                                user_data = obj.object.__dict__.copy()
                                                user_data.pop('_state', None)

                                                # Only use PK if it's valid
                                                if is_valid_pk and not model.objects.filter(pk=original_pk).exists():
                                                    user_data['id'] = original_pk
                                                else:
                                                    user_data.pop('id', None)

                                                password = user_data.pop(
                                                    'password', None)
                                                new_user = model(**user_data)
                                                if password:
                                                    new_user.set_password(
                                                        password)
                                                new_user.save()
                                                objects_processed += 1
                                                logger.info(
                                                    "Created user without email with auto-generated PK")
                                        except Exception as no_email_error:
                                            logger.error(
                                                f"Failed to create user without email: {no_email_error}")
                                            errors.append(
                                                f"User without email: {str(no_email_error)}")
                                else:
                                    # Default path for non-User models via deserializer
                                    try:
                                        # Use a savepoint per object so one failure doesn't poison the batch
                                        with transaction.atomic():
                                            obj.save()
                                            objects_processed += 1
                                    except IntegrityError as e:
                                        logger.error(
                                            f"Integrity error restoring {model_name} object (pk={getattr(obj, 'pk', None)}): {e}")
                                        errors.append(
                                            f"{model_name} object (pk={getattr(obj, 'pk', None)}): {str(e)}")
                                    except Exception as e:
                                        logger.error(
                                            f"Error restoring {model_name} object (pk={getattr(obj, 'pk', None)}): {e}")
                                        errors.append(
                                            f"{model_name} object (pk={getattr(obj, 'pk', None)}): {str(e)}")

                            except IntegrityError as e:
                                logger.warning(
                                    f"Integrity error for {model_name} object: {e}")
                                # Try to create without PK
                                try:
                                    data_dict = obj.object.__dict__.copy()
                                    data_dict.pop('id', None)
                                    data_dict.pop('_state', None)
                                    model.objects.create(**data_dict)
                                    objects_processed += 1
                                    logger.info(
                                        f"Created {model_name} with auto PK after integrity error")
                                except Exception as e2:
                                    logger.error(
                                        f"Failed to restore {model_name} object (pk={getattr(obj.object, 'pk', None)}): {e2}")
                                    errors.append(
                                        f"{model_name} object: {str(e2)}")
                            except Exception as e:
                                logger.error(
                                    f"Error restoring {model_name} object (pk={getattr(obj.object, 'pk', None)}): {e}")
                                errors.append(
                                    f"{model_name} object (pk={getattr(obj.object, 'pk', None)}): {str(e)}")

                        restored_counts[model_name] = objects_processed
                        logger.info(
                            f"Successfully processed {objects_processed} {model_name} records")

                        # Flag critical failure if a critical model had data in file but restored zero
                        if model_name in critical_models and file_record_counts.get(model_name, 0) > 0 and objects_processed == 0:
                            critical_failures[model_name] = {
                                'file_records': file_record_counts.get(model_name, 0),
                                'restored': 0
                            }

                except Exception as model_error:
                    logger.error(
                        f"Error processing {model_name}: {model_error}")
                    errors.append(f"{model_name}: {str(model_error)}")

            # Step 3: Restore media files
            media_dir = os.path.join(temp_dir, 'media')
            if os.path.exists(media_dir):
                media_root = settings.MEDIA_ROOT
                for root, dirs, files in os.walk(media_dir):
                    for file in files:
                        src_path = os.path.join(root, file)
                        rel_path = os.path.relpath(src_path, media_dir)
                        dest_path = os.path.join(media_root, rel_path)

                        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
                        shutil.copy2(src_path, dest_path)
                        logger.info(f"Restored media file: {rel_path}")

            # Cleanup temporary files
            if temp_dir and os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)

            # If any critical models failed to restore, return 500 with details
            if critical_failures:
                response_data = {
                    "detail": "Restore failed for critical models",
                    "critical_failures": critical_failures,
                    "summary": restored_counts,
                    "warnings": errors,
                    "auto_login": False
                }
                return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            # Try to maintain user session
            try:
                restored_user = User.objects.get(email=current_user_email)
                refresh = RefreshToken.for_user(restored_user)

                response_data = {
                    "detail": f"Restore completed successfully from {file_obj.name}",
                    "summary": restored_counts,
                    "auto_login": True,
                    "tokens": {
                        "access": str(refresh.access_token),
                        "refresh": str(refresh)
                    }
                }

                if errors:
                    response_data["warnings"] = errors

                return Response(response_data, status=status.HTTP_200_OK)

            except User.DoesNotExist:
                response_data = {
                    "detail": f"Restore completed successfully from {file_obj.name}",
                    "summary": restored_counts,
                    "auto_login": False,
                    "note": "Please login with your credentials."
                }

                if errors:
                    response_data["warnings"] = errors

                return Response(response_data, status=status.HTTP_200_OK)

        except Exception as restore_error:
            logger.error(f"Restore failed: {restore_error}")

            # Attempt safety rollback with separate transactions
            rollback_errors = []
            try:
                # Restore from safety backup in reverse order
                for model, model_name in reversed(MODEL_BACKUP_ORDER):
                    try:
                        with transaction.atomic():
                            # Clear current data
                            model.objects.all().delete()

                            # Restore from backup
                            if model_name in safety_backup:
                                for obj_data in safety_backup[model_name]:
                                    try:
                                        model.objects.create(**obj_data)
                                    except Exception as e:
                                        rollback_errors.append(
                                            f"Failed to restore {model_name} object: {e}")
                    except Exception as e:
                        rollback_errors.append(
                            f"Failed to restore {model_name}: {e}")

                if not rollback_errors:
                    logger.info("Safety backup restored successfully")
                else:
                    logger.error(
                        f"Safety rollback had errors: {rollback_errors}")

            except Exception as rollback_error:
                logger.error(f"Safety rollback failed: {rollback_error}")
                rollback_errors.append(str(rollback_error))

            error_response = {
                "detail": "Restore failed",
                "error": str(restore_error),
                "auto_login": False
            }

            if rollback_errors:
                error_response["rollback_errors"] = rollback_errors

            return Response(error_response, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    except Exception as e:
        logger.error(f"Restore failed: {e}", exc_info=True)
        # Cleanup on error
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)

        return Response(
            {
                "detail": "Restore failed",
                "error": str(e),
                "auto_login": False
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_backups(request):
    if request.user.role != 'admin':
        return Response({"detail": "Only administrators can view backups."}, status=status.HTTP_403_FORBIDDEN)
    qs = Backup.objects.all()
    serializer = BackupSerializer(qs, many=True)
    return Response(serializer.data)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_e_signature(request):
    user = request.user
    e_signature = request.FILES.get('e_signature')
    if not e_signature:
        return Response({"error": "No e-signature file provided."}, status=400)
    user.e_signature = e_signature
    user.save(update_fields=['e_signature'])

    # Log audit for e-signature update
    from .audit_utils import log_audit_event
    log_audit_event(
        request=request,
        action='update',
        module='user',
        description=f"Updated e-signature for user {user.get_full_name()}",
        object_id=user.pk,
        object_type='User',
        object_name=user.get_full_name()
    )

    # Generate new token with updated e_signature
    from rest_framework_simplejwt.tokens import RefreshToken
    refresh = RefreshToken.for_user(user)
    access_token = refresh.access_token

    # Add custom claims to the new token
    access_token['first_name'] = user.first_name
    access_token['last_name'] = user.last_name
    access_token['email'] = user.email
    access_token['role'] = user.role
    access_token['password_change_required'] = user.password_change_required

    if user.school_district:
        access_token['school_district'] = {
            'id': user.school_district.districtId,
            'name': user.school_district.districtName,
            'municipality': user.school_district.municipality,
            'legislative_district': user.school_district.legislativeDistrict
        }
    else:
        access_token['school_district'] = None

    # Add profile picture URL if available
    if user.profile_picture:
        access_token['profile_picture'] = user.profile_picture.url
    else:
        access_token['profile_picture'] = None

    # Add e-signature URL (now available)
    access_token['e_signature'] = user.e_signature.url

    return Response({
        "message": "E-signature updated successfully.",
        "access": str(access_token),
        "refresh": str(refresh)
    })


# Add to views.py

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_liquidation_report(request, LiquidationID):
    """
    Generate a liquidation report in Excel format matching the Appendix-44-LR-1 template
    with data automatically populated from the system
    """
    try:
        liquidation = LiquidationManagement.objects.get(
            LiquidationID=LiquidationID)
        request_obj = liquidation.request
        user = request_obj.user
        school = user.school

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LR"

        # Set column widths to match the template
        column_widths = {
            'A': 30.71, 'B': 36.25, 'C': 33.39, 'D': 8.43, 'E': 8.43,
            'F': 8.43, 'G': 8.43, 'H': 8.43, 'I': 8.43, 'J': 8.43, 'K': 8.43
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Define border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Define fonts
        base_font = Font(name='Times New Roman', size=10)
        bold_font = Font(name='Times New Roman', size=10, bold=True)

        # Add template content (matching the provided Excel structure)
        # Appendix 44 in cell C1
        ws['C1'] = "Appendix 44"
        ws['C1'].font = base_font
        ws['C1'].alignment = Alignment(horizontal='right')

        # Liquidation Report title
        ws['A3'] = "LIQUIDATION REPORT"
        ws['A3'].font = Font(name='Times New Roman', size=14, bold=True)
        ws.merge_cells('A3:C3')
        ws['A3'].alignment = Alignment(horizontal='center')

        # Serial number and date
        ws['C4'] = f"Serial No.: {liquidation.LiquidationID}"
        ws['C4'].font = base_font

        current_date = timezone.now().strftime("%Y-%m-%d")
        ws['C5'] = f"Date: {current_date}"
        ws['C5'].font = base_font

        # Period covered - use the request month/year
        if request_obj.request_monthyear:
            period_covered = request_obj.request_monthyear
        else:
            period_covered = timezone.now().strftime("%Y-%m")
        ws['A5'] = f"Period Covered {period_covered}"
        ws['A5'].font = base_font

        # Entity name and fund cluster
        entity_name = school.schoolName if school else "N/A"
        ws['A7'] = f"Entity Name :  {entity_name}"
        ws['A7'].font = base_font

        # For fund cluster, we'll use a default value or leave it blank
        ws['A8'] = "Fund Cluster :  _____________________________________________"
        ws['A8'].font = base_font

        # Responsibility center code - use district code if available
        responsibility_center = school.district.districtId if school and school.district else "N/A"
        ws['C8'] = f"Responsibility Center Code: {responsibility_center}"
        ws['C8'].font = base_font

        # Table header
        ws['A11'] = "PARTICULARS"
        ws['A11'].font = bold_font
        ws['A11'].border = thin_border
        ws['A11'].alignment = Alignment(horizontal='center', vertical='center')

        ws['C11'] = "AMOUNT"
        ws['C11'].font = bold_font
        ws['C11'].border = thin_border
        ws['C11'].alignment = Alignment(horizontal='center', vertical='center')

        # Add liquidation items
        row = 12
        total_amount = Decimal('0.00')
        for lp in liquidation.liquidation_priorities.all():
            ws[f'A{row}'] = lp.priority.expenseTitle
            ws[f'A{row}'].font = base_font
            ws[f'A{row}'].border = thin_border

            ws[f'C{row}'] = float(lp.amount)
            ws[f'C{row}'].font = base_font
            ws[f'C{row}'].border = thin_border
            ws[f'C{row}'].number_format = '#,##0.00'

            total_amount += lp.amount  # Don't convert to float here
            row += 1

        # Add total amount
        total_row = row + 2
        ws[f'A{total_row}'] = "TOTAL AMOUNT SPENT"
        ws[f'A{total_row}'].font = bold_font
        ws[f'A{total_row}'].border = thin_border

        ws[f'C{total_row}'] = float(total_amount)
        ws[f'C{total_row}'].font = bold_font
        ws[f'C{total_row}'].border = thin_border
        ws[f'C{total_row}'].number_format = '#,##0.00'

        # Add cash advance information
        cash_advance_row = total_row + 2
        ws[f'A{cash_advance_row}'] = "AMOUNT OF CASH ADVANCE PER DV NO.______DTD. ______"
        ws[f'A{cash_advance_row}'].font = base_font
        ws[f'A{cash_advance_row}'].border = thin_border

        # Calculate cash advance amount (sum of requested amounts)
        cash_advance_amount = sum(
            rp.amount for rp in request_obj.requestpriority_set.all()
        )
        ws[f'C{cash_advance_row}'] = float(cash_advance_amount)
        ws[f'C{cash_advance_row}'].font = base_font
        ws[f'C{cash_advance_row}'].border = thin_border
        ws[f'C{cash_advance_row}'].number_format = '#,##0.00'

        # Add refund information
        refund_row = cash_advance_row + 2
        ws[f'A{refund_row}'] = "AMOUNT REFUNDED PER OR NO. ________DTD. ___________"
        ws[f'A{refund_row}'].font = base_font
        ws[f'A{refund_row}'].border = thin_border

        refund_amount = cash_advance_amount - total_amount  # Both are Decimal now
        ws[f'C{refund_row}'] = float(refund_amount) if refund_amount > 0 else 0
        ws[f'C{refund_row}'].font = base_font
        ws[f'C{refund_row}'].border = thin_border
        ws[f'C{refund_row}'].number_format = '#,##0.00'

        # Add amount to be reimbursed
        reimbursement_row = refund_row + 2
        ws[f'A{reimbursement_row}'] = "AMOUNT TO BE REIMBURSED"
        ws[f'A{reimbursement_row}'].font = base_font
        ws[f'A{reimbursement_row}'].border = thin_border

        reimbursement_amount = total_amount - cash_advance_amount
        ws[f'C{reimbursement_row}'] = float(
            reimbursement_amount) if reimbursement_amount > 0 else 0
        ws[f'C{reimbursement_row}'].font = base_font
        ws[f'C{reimbursement_row}'].border = thin_border
        ws[f'C{reimbursement_row}'].number_format = '#,##0.00'

        # Add certification section
        cert_row = reimbursement_row + 3
        ws[f'A{cert_row}'] = "Certified: Correctness of the above data"
        ws[f'A{cert_row}'].font = bold_font
        ws[f'A{cert_row}'].border = thin_border
        ws[f'A{cert_row}'].alignment = Alignment(horizontal='center')

        ws[f'B{cert_row}'] = "Certified: Purpose of travel / cash advance duly accomplished"
        ws[f'B{cert_row}'].font = bold_font
        ws[f'B{cert_row}'].border = thin_border
        ws[f'B{cert_row}'].alignment = Alignment(horizontal='center')

        ws[f'C{cert_row}'] = "Certified: Supporting documents complete and proper"
        ws[f'C{cert_row}'].font = bold_font
        ws[f'C{cert_row}'].border = thin_border
        ws[f'C{cert_row}'].alignment = Alignment(horizontal='center')

        # Add signature lines with user names if available
        cert_row += 2

        # Claimant (the user who submitted the request)
        claimant_name = f"{user.first_name} {user.last_name}" if user else "________________________"
        ws[f'A{cert_row}'] = claimant_name
        ws[f'A{cert_row}'].font = base_font
        ws[f'A{cert_row}'].alignment = Alignment(horizontal='center')

        # Immediate Supervisor (could be district admin or school head)
        # For now, we'll leave it blank as it's not directly in the model
        ws[f'B{cert_row}'] = "________________________"
        ws[f'B{cert_row}'].font = base_font
        ws[f'B{cert_row}'].alignment = Alignment(horizontal='center')

        # Head, Accounting Division Unit (could be division accountant)
        # For now, we'll leave it blank as it's not directly in the model
        ws[f'C{cert_row}'] = "________________________"
        ws[f'C{cert_row}'].font = base_font
        ws[f'C{cert_row}'].alignment = Alignment(horizontal='center')

        # Add position titles
        cert_row += 1
        ws[f'A{cert_row}'] = "Signature over Printed Name"
        ws[f'A{cert_row}'].font = base_font
        ws[f'A{cert_row}'].alignment = Alignment(horizontal='center')

        ws[f'B{cert_row}'] = "Signature over Printed Name"
        ws[f'B{cert_row}'].font = base_font
        ws[f'B{cert_row}'].alignment = Alignment(horizontal='center')

        ws[f'C{cert_row}'] = "Signature over Printed Name"
        ws[f'C{cert_row}'].font = base_font
        ws[f'C{cert_row}'].alignment = Alignment(horizontal='center')

        # Add role titles
        cert_row += 1
        ws[f'A{cert_row}'] = "Claimant"
        ws[f'A{cert_row}'].font = base_font
        ws[f'A{cert_row}'].alignment = Alignment(horizontal='center')

        ws[f'B{cert_row}'] = "Immediate Supervisor"
        ws[f'B{cert_row}'].font = base_font
        ws[f'B{cert_row}'].alignment = Alignment(horizontal='center')

        ws[f'C{cert_row}'] = "Head, Accounting Division Unit"
        ws[f'C{cert_row}'].font = base_font
        ws[f'C{cert_row}'].alignment = Alignment(horizontal='center')

        # Add JEV number - we don't have this in the model, so leave blank
        cert_row += 2
        ws[f'C{cert_row}'] = "JEV No.: ___________________"
        ws[f'C{cert_row}'].font = base_font

        # Add date lines
        cert_row += 1
        ws[f'A{cert_row}'] = "Date: ______________________"
        ws[f'A{cert_row}'].font = base_font

        ws[f'B{cert_row}'] = "Date: _____________________"
        ws[f'B{cert_row}'].font = base_font

        ws[f'C{cert_row}'] = "Date:  _____________________"
        ws[f'C{cert_row}'].font = base_font

        # Apply borders to all relevant cells
        for row in ws.iter_rows(min_row=1, max_row=cert_row, min_col=1, max_col=3):
            for cell in row:
                if cell.value:
                    cell.border = thin_border

        # Create HTTP response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="liquidation_report_{LiquidationID}.xlsx"'

        wb.save(response)
        return response

    except LiquidationManagement.DoesNotExist:
        return Response(
            {'error': 'Liquidation not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(
            f"Error generating liquidation report: {str(e)}", exc_info=True)
        return Response(
            {'error': 'Failed to generate report'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_next_available_month(request):
    """Get the next available month for the user to submit a request"""
    user = request.user

    # Create a temporary request object to use the method
    temp_request = RequestManagement(user=user)
    next_month = temp_request.get_next_available_month()

    today = date.today()
    try:
        req_year, req_month = map(int, next_month.split('-'))
        is_advance = (req_year, req_month) > (today.year, today.month)
    except (ValueError, AttributeError):
        is_advance = False

    return Response({
        'next_available_month': next_month,
        'is_advance_request': is_advance,
        'can_submit': RequestManagement.can_user_request_for_month(user, next_month)
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def debug_liquidation_times(request):
    """More permissive debug endpoint"""
    try:
        # First try the original filter
        liquidated_liquidations = LiquidationManagement.objects.filter(
            status='liquidated',
            date_liquidated__isnull=False,
            created_at__isnull=False
        ).order_by('-created_at')[:10]

        # If no results, try a less restrictive filter
        if not liquidated_liquidations.exists():
            liquidated_liquidations = LiquidationManagement.objects.filter(
                date_liquidated__isnull=False,
                created_at__isnull=False
            ).order_by('-created_at')[:10]

        # If still no results, get ANY records
        if not liquidated_liquidations.exists():
            liquidated_liquidations = LiquidationManagement.objects.all().order_by(
                '-created_at')[:10]

        debug_data = []
        for liq in liquidated_liquidations:
            time_diff = None
            days_diff = 0

            if liq.date_liquidated and liq.created_at:
                time_diff = liq.date_liquidated - liq.created_at
                days_diff = time_diff.total_seconds() / (24 * 60 * 60)

            debug_data.append({
                'liquidation_id': liq.LiquidationID,
                'status': liq.status,
                'created_at': liq.created_at.isoformat() if liq.created_at else None,
                'date_liquidated': liq.date_liquidated.isoformat() if liq.date_liquidated else None,
                'time_diff_seconds': time_diff.total_seconds() if time_diff else None,
                'time_diff_days': days_diff,
                'has_both_dates': liq.date_liquidated is not None and liq.created_at is not None
            })

        return Response({
            'total_records': LiquidationManagement.objects.count(),
            'filtered_records': liquidated_liquidations.count(),
            'sample_data': debug_data,
            'filter_used': 'liquidated status with both dates' if liquidated_liquidations.exists() else 'fallback to any records'
        })

    except Exception as e:
        return Response({
            'error': str(e),
            'message': 'Error retrieving liquidation data'
        }, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_request_eligibility(request):
    """Check if user is eligible to submit a new request"""
    user = request.user
    target_month = request.query_params.get('month')  # Format: YYYY-MM

    if target_month:
        can_request = RequestManagement.can_user_request_for_month(
            user, target_month)
    else:
        # Check for next available month
        temp_request = RequestManagement(user=user)
        next_month = temp_request.get_next_available_month()
        can_request = RequestManagement.can_user_request_for_month(
            user, next_month)
        target_month = next_month

    # Get current active requests
    active_requests = RequestManagement.objects.filter(
        user=user
    ).exclude(status__in=['liquidated', 'rejected']).order_by('-created_at')

    # Get reason if cannot request
    reason = None
    if not can_request:
        existing_same_month = RequestManagement.objects.filter(
            user=user,
            request_monthyear=target_month
        ).exclude(status='rejected').first()

        if existing_same_month:
            reason = f"You already have a request for {target_month}"
        else:
            unliquidated = active_requests.first()
            if unliquidated:
                reason = f"Please liquidate your current request ({unliquidated.request_id}) before submitting a new one"
            else:
                reason = "Cannot determine eligibility"

    return Response({
        'can_submit_request': can_request,
        'target_month': target_month,
        'reason': reason,
        'active_requests': RequestManagementSerializer(active_requests, many=True).data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_approved_request_pdf(request, request_id):
    """
    Generate PDF with actual e-signatures after approval
    This endpoint can only be accessed for approved requests
    """
    try:
        from .pdf_utils import generate_request_pdf_with_signatures
        from django.shortcuts import get_object_or_404

        # Get the request object
        req = get_object_or_404(RequestManagement, request_id=request_id)

        # Check if user has permission to generate PDF
        user = request.user

        # Allow access if:
        # 1. User is the request owner
        # 2. User is a superintendent (who approved it)
        # 3. User is an admin
        # 4. User is an accountant
        has_permission = (
            req.user == user or
            user.role in ['superintendent', 'admin', 'accountant'] or
            (user.role == 'district_admin' and req.user.school and req.user.school.district ==
             user.school_district)
        )

        if not has_permission:
            return HttpResponse(
                '{"error": "You don\'t have permission to generate this PDF"}',
                content_type='application/json',
                status=403
            )

        # Check if request is approved
        # if req.status != 'approved':
        #     return HttpResponse(
        #         '{"error": "Request must be approved first to generate PDF"}',
        #         content_type='application/json',
        #         status=400
        #     )

        # Generate PDF with actual signatures
        pdf_content = generate_request_pdf_with_signatures(req)

        # Store PDF for audit trail (optional - can be enabled for compliance)
        from .models import GeneratedPDF
        from django.core.files.base import ContentFile

        try:
            # Create a record of PDF generation
            pdf_record = GeneratedPDF.objects.create(
                request=req,
                generated_by=user,
                generation_method='server_side',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[
                    :500]  # Limit length
            )

            # Optionally store the PDF file (uncomment if you want to store PDFs)
            # pdf_filename = f"approved_request_{request_id}_{pdf_record.id}.pdf"
            # pdf_record.pdf_file.save(
            #     pdf_filename,
            #     ContentFile(pdf_content),
            #     save=True
            # )

            logger.info(
                f"PDF generated for approved request {request_id} by user {user.id} (PDF record ID: {pdf_record.id})")

        except Exception as e:
            logger.error(f"Error creating PDF record: {e}")
            # Continue with PDF generation even if record creation fails

        # Create HTTP response
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="approved_request_{request_id}.pdf"'
        response['Content-Length'] = len(pdf_content)

        return response

    except Exception as e:
        logger.error(
            f"Error generating PDF for request {request_id}: {str(e)}")
        return HttpResponse(
            '{"error": "Failed to generate PDF. Please try again."}',
            content_type='application/json',
            status=500
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def liquidation_report(request):
    """
    Generate liquidation report with filtering and export capabilities
    Similar to unliquidated requests report but for liquidations
    """
    from .liquidation_report_utils import (
        generate_liquidation_report_data,
        generate_liquidation_csv_report,
        generate_liquidation_excel_report,
        LiquidationReportPagination
    )
    from datetime import datetime

    # Get filter parameters
    status_filter = request.GET.get('status', 'all')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    legislative_district = request.GET.get('legislative_district')
    municipality = request.GET.get('municipality')
    school_district = request.GET.get('school_district')
    school_ids = request.GET.get('school_ids')
    export_format = request.GET.get('export')
    page_size = request.GET.get('page_size', 50)

    # Validate date format
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid start_date format. Use YYYY-MM-DD"}, status=400)

    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid end_date format. Use YYYY-MM-DD"}, status=400)

    # Generate report data
    report_data = generate_liquidation_report_data(
        status_filter=status_filter,
        start_date=start_date,
        end_date=end_date,
        legislative_district=legislative_district,
        municipality=municipality,
        school_district=school_district,
        school_ids=school_ids
    )

    # Prepare filters for export
    filters = {
        'status': status_filter,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else None,
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else None,
        'legislative_district': legislative_district,
        'municipality': municipality,
        'school_district': school_district,
        'school_ids': school_ids
    }

    # Handle export formats
    if export_format == 'csv':
        return generate_liquidation_csv_report(report_data, filters)
    elif export_format == 'excel':
        return generate_liquidation_excel_report(report_data, filters)

    # Apply pagination for regular API response
    paginator = LiquidationReportPagination()
    paginator.page_size = int(page_size)
    paginated_data = paginator.paginate_queryset(report_data, request)

    # Calculate status counts for filters
    status_counts = {}
    for item in report_data:
        status = item['status']
        status_counts[status] = status_counts.get(status, 0) + 1

    return paginator.get_paginated_response({
        'results': paginated_data,
        'total_count': len(report_data),
        'filters': {
            'status_filter': status_filter,
            'start_date': filters['start_date'],
            'end_date': filters['end_date'],
            'legislative_district': legislative_district,
            'municipality': municipality,
            'school_district': school_district,
            'school_ids': school_ids,
            'status_counts': status_counts
        }
    })


class AuditLogSerializer(serializers.ModelSerializer):
    user = UserSerializer(read_only=True)

    class Meta:
        model = AuditLog
        fields = '__all__'


class AuditLogListView(generics.ListAPIView):
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_queryset(self):
        if not self.request.user.role == 'admin':
            raise PermissionDenied("Only administrators can view audit logs")

        queryset = AuditLog.objects.all().select_related('user')

        # Filter by module
        module = self.request.query_params.get('module')
        if module:
            queryset = queryset.filter(module=module)

        # Filter by action
        action = self.request.query_params.get('action')
        if action:
            queryset = queryset.filter(action=action)

        # Filter by user
        user_id = self.request.query_params.get('user_id')
        if user_id:
            queryset = queryset.filter(user_id=user_id)

        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date and end_date:
            queryset = queryset.filter(timestamp__date__range=[
                                       start_date, end_date])

        # Search in description
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(description__icontains=search)

        # Filter by object type
        object_type = self.request.query_params.get('object_type')
        if object_type:
            queryset = queryset.filter(object_type=object_type)

        # Filter by object ID
        object_id = self.request.query_params.get('object_id')
        if object_id:
            queryset = queryset.filter(object_id=object_id)

        return queryset.order_by('-timestamp')

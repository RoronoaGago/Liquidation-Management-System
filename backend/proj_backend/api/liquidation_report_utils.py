# liquidation_report_utils.py
import io
import csv
import os
from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone
from rest_framework.pagination import PageNumberPagination
from .models import LiquidationManagement, RequestManagement
from django.db.models import Q, Sum
import logging
from datetime import date, timedelta

logger = logging.getLogger('api')


class LiquidationReportPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 1000


def generate_liquidation_report_data(
    status_filter=None,
    start_date=None,
    end_date=None,
    legislative_district=None,
    municipality=None,
    school_district=None,
    school_ids=None
):
    """Generate liquidation report data with filtering"""

    # Start with all liquidations
    queryset = LiquidationManagement.objects.select_related(
        'request__user__school__district'
    ).all()

    # Apply status filter
    if status_filter and status_filter != 'all':
        queryset = queryset.filter(status=status_filter)

    # Apply date range filter - FIXED: Use timezone-aware date comparison
    if start_date and end_date:
        # Convert string dates to timezone-aware datetime objects
        start_datetime = timezone.make_aware(
            timezone.datetime.combine(start_date, timezone.datetime.min.time()))
        end_datetime = timezone.make_aware(
            timezone.datetime.combine(end_date, timezone.datetime.max.time()))

        queryset = queryset.filter(
            created_at__gte=start_datetime,
            created_at__lte=end_datetime
        )
    elif start_date:
        start_datetime = timezone.make_aware(
            timezone.datetime.combine(start_date, timezone.datetime.min.time()))
        queryset = queryset.filter(created_at__gte=start_datetime)
    elif end_date:
        end_datetime = timezone.make_aware(
            timezone.datetime.combine(end_date, timezone.datetime.max.time()))
        queryset = queryset.filter(created_at__lte=end_datetime)

    # Apply legislative district filter
    if legislative_district:
        queryset = queryset.filter(
            request__user__school__district__legislativeDistrict=legislative_district
        )

    # Apply municipality filter
    if municipality:
        queryset = queryset.filter(
            request__user__school__district__municipality=municipality
        )

    # Apply school district filter
    if school_district:
        queryset = queryset.filter(
            request__user__school__district__districtId=school_district
        )

    # Apply school IDs filter
    if school_ids:
        school_id_list = school_ids.split(',') if isinstance(
            school_ids, str) else school_ids
        queryset = queryset.filter(
            request__user__school__schoolId__in=school_id_list
        )

    # Convert to list of dictionaries - FIXED: Convert User objects to strings
    report_data = []
    for liquidation in queryset:
        request = liquidation.request
        user = request.user
        school = user.school
        district = school.district if school else None

        # Helper function to format user information
        def format_user(user_obj):
            if user_obj:
                return f"{user_obj.first_name} {user_obj.last_name} ({user_obj.email})"
            return ""

        report_data.append({
            'liquidation_id': liquidation.LiquidationID,
            'school_id': school.schoolId if school else '',
            'school_name': school.schoolName if school else '',
            'district_name': district.districtName if district else '',
            'municipality': district.municipality if district else '',
            'legislative_district': district.legislativeDistrict if district else '',
            'created_at': liquidation.created_at,
            'request_month': request.request_monthyear,
            'status': liquidation.status,
            'date_submitted': liquidation.date_submitted,
            'date_liquidated': liquidation.date_liquidated,
        })

    return report_data


def generate_liquidation_csv_report(report_data, filters):
    """Generate CSV report for liquidation data"""
    response = HttpResponse(content_type='text/csv')

    # Generate filename based on filters
    filename_parts = ['liquidation_report']
    if filters.get('start_date') and filters.get('end_date'):
        filename_parts.append(
            f"{filters['start_date']}_to_{filters['end_date']}")
    elif filters.get('status') and filters['status'] != 'all':
        filename_parts.append(filters['status'])

    response['Content-Disposition'] = f'attachment; filename="{"_".join(filename_parts)}.csv"'

    writer = csv.writer(response)

    # Add company header
    writer.writerow(['COMPANY NAME'])
    writer.writerow(['Liquidation Report'])
    writer.writerow([f'Generated on: {timezone.now().date()}'])
    if filters.get('start_date') and filters.get('end_date'):
        writer.writerow(
            [f'Date Range: {filters["start_date"]} to {filters["end_date"]}'])
    if filters.get('status') and filters['status'] != 'all':
        writer.writerow([f'Status: {filters["status"]}'])
    writer.writerow([])  # Empty row for spacing

    # Write column headers (removed unwanted columns)
    writer.writerow([
        'Liquidation ID', 'School ID', 'School Name',
        'District', 'Municipality', 'Legislative District', 'Month',
        'Status', 'Submitted At', 'Liquidated At'
    ])

    for item in report_data:
        writer.writerow([
            item['liquidation_id'],
            item['school_id'],
            item['school_name'],
            item['district_name'],
            item['municipality'],
            item['legislative_district'],
            item['request_month'],
            item['status'],
            item['date_submitted'].date() if item['date_submitted'] else 'N/A',
            item['date_liquidated'].date() if item['date_liquidated'] else 'N/A',
        ])

    return response


def get_liquidation_summary_stats(
    start_date=None,
    end_date=None,
    legislative_district=None,
    municipality=None,
    school_district=None,
    school_ids=None
):
    """Get summary statistics for liquidation dashboard"""

    # Start with all liquidations
    queryset = LiquidationManagement.objects.select_related(
        'request__user__school__district'
    ).all()

    # Apply date range filter - FIXED: Handle string dates properly
    if start_date and end_date:
        # Convert string dates to datetime.date objects first
        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Now convert to timezone-aware datetime objects
        start_datetime = timezone.make_aware(
            datetime.combine(start_date, timezone.datetime.min.time()))
        end_datetime = timezone.make_aware(
            datetime.combine(end_date, timezone.datetime.max.time()))

        queryset = queryset.filter(
            created_at__gte=start_datetime,
            created_at__lte=end_datetime
        )
    elif start_date:
        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        start_datetime = timezone.make_aware(
            datetime.combine(start_date, timezone.datetime.min.time()))
        queryset = queryset.filter(created_at__gte=start_datetime)
    elif end_date:
        if isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        end_datetime = timezone.make_aware(
            datetime.combine(end_date, timezone.datetime.max.time()))
        queryset = queryset.filter(created_at__lte=end_datetime)

    # Apply legislative district filter
    if legislative_district:
        queryset = queryset.filter(
            request__user__school__district__legislativeDistrict=legislative_district
        )

    # Apply municipality filter
    if municipality:
        queryset = queryset.filter(
            request__user__school__district__municipality=municipality
        )

    # Apply school district filter
    if school_district:
        queryset = queryset.filter(
            request__user__school__district__districtId=school_district
        )

    # Apply school IDs filter
    if school_ids:
        school_id_list = school_ids.split(',') if isinstance(
            school_ids, str) else school_ids
        queryset = queryset.filter(
            request__user__school__schoolId__in=school_id_list
        )

    # Calculate summary statistics
    total_liquidations = queryset.count()

    # Liquidated status
    liquidated_count = queryset.filter(status='liquidated').count()

    # Pending review statuses (all review-related statuses)
    pending_review_statuses = [
        'submitted',
        'under_review_district',
        'approved_district',
        'under_review_liquidator',
        'approved_liquidator',
        'under_review_division'
    ]
    pending_review_count = queryset.filter(
        status__in=pending_review_statuses).count()

    # Needs revision
    needs_revision_count = queryset.filter(status='resubmit').count()

    # Draft status
    draft_count = queryset.filter(status='draft').count()

    return {
        'total_liquidations': total_liquidations,
        'liquidated': liquidated_count,
        'pending_review': pending_review_count,
        'needs_revision': needs_revision_count,
        'draft': draft_count
    }


def generate_liquidation_excel_report(report_data, filters, request_user=None):
    """Generate Excel report for liquidation data with summary at bottom"""
    # Get the active accountant for signature
    from .models import User
    active_accountant = User.objects.filter(
        role='accountant', 
        is_active=True
    ).first()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidation Report"

    # Define headers for the report
    headers = [
        'Liquidation ID', 'School ID', 'School Name',
        'District', 'Municipality', 'Legislative District', 'Month',
        'Status', 'Submitted At', 'Liquidated At'
    ]

    # Add logo
    try:
        logo_path = os.path.join(
            settings.BASE_DIR, 'static', 'images', 'logo.png')
        if os.path.exists(logo_path):
            logo_img = Image(logo_path)
            logo_img.width = 110
            logo_img.height = 96
            ws.add_image(logo_img, 'B2')
    except Exception as e:
        logger.error(f"Could not add logo: {e}")

    # Define fonts and styles
    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=12, bold=True)
    data_font = Font(name='Calibri', size=11)
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    # Define border
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Add report header information
    ws['D3'] = "Department of Education"
    ws['D3'].font = title_font
    ws['D3'].alignment = center_alignment
    ws.merge_cells('D3:H3')

    ws['D4'] = "La Union Schools Division Office"
    ws['D4'].font = Font(name='Calibri', size=14, bold=True)
    ws['D4'].alignment = center_alignment
    ws.merge_cells('D4:H4')

    # Add date range to title if available
    report_title = "LIQUIDATION LIST REPORT"
    if filters.get('start_date') and filters.get('end_date'):
        report_title += f" for {filters['start_date']} to {filters['end_date']}"

    ws.merge_cells('C5:J5')
    ws['C5'] = report_title
    ws['C5'].font = title_font
    ws['C5'].alignment = center_alignment

    ws['D6'] = f"Date Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['D6'].font = Font(name='Calibri', size=11)
    ws['D6'].alignment = center_alignment
    ws.merge_cells('D6:H6')

    # Add filter information (excluding start/end dates as they're already in the header)
    filter_row = 8
    ws[f'A{filter_row}'] = "APPLIED FILTERS"
    ws[f'A{filter_row}'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{filter_row}:D{filter_row}')

    filter_info = []
    if filters.get('status') and filters['status'] != 'all':
        filter_info.append(f"Status: {filters['status']}")
    if filters.get('legislative_district'):
        filter_info.append(
            f"Legislative District: {filters['legislative_district']}")
    if filters.get('municipality'):
        filter_info.append(f"Municipality: {filters['municipality']}")
    if filters.get('school_district'):
        filter_info.append(f"School District: {filters['school_district']}")
    if filters.get('school_ids'):
        filter_info.append(f"School IDs: {filters['school_ids']}")

    if not filter_info:
        filter_info.append("No additional filters applied")

    for i, filter_text in enumerate(filter_info, start=filter_row + 1):
        ws[f'A{i}'] = filter_text
        ws[f'A{i}'].font = Font(name='Calibri', size=10)
        ws.merge_cells(f'A{i}:D{i}')

    # Add column headers (starting after filters with proper spacing)
    header_row = filter_row + len(filter_info) + 2
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        # Add light gray background to header
        cell.fill = PatternFill(start_color="F0F0F0",
                                end_color="F0F0F0", fill_type="solid")

    # Set header row height
    ws.row_dimensions[header_row].height = 25

    # Add data rows
    data_start_row = header_row + 1
    for row_num, item in enumerate(report_data, data_start_row):
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Set data row height
        ws.row_dimensions[row_num].height = 20

        # Fill in data
        ws.cell(row=row_num, column=1, value=item['liquidation_id'])
        ws.cell(row=row_num, column=2, value=item['school_id'])
        ws.cell(row=row_num, column=3, value=item['school_name'])
        ws.cell(row=row_num, column=4, value=item['district_name'])
        ws.cell(row=row_num, column=5, value=item['municipality'])
        ws.cell(row=row_num, column=6, value=item['legislative_district'])
        ws.cell(row=row_num, column=7, value=item['request_month'])
        ws.cell(row=row_num, column=8, value=item['status'])
        ws.cell(row=row_num, column=9, value=item['date_submitted'].date(
        ) if item['date_submitted'] else 'N/A')
        ws.cell(row=row_num, column=10, value=item['date_liquidated'].date(
        ) if item['date_liquidated'] else 'N/A')

    # ========== SUMMARY STATISTICS AT THE BOTTOM ==========
    summary_start_row = data_start_row + len(report_data) + 3  # +3 for spacing

    # Summary header
    ws[f'A{summary_start_row}'] = "SUMMARY STATISTICS"
    ws[f'A{summary_start_row}'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{summary_start_row}:D{summary_start_row}')

    # Get summary statistics
    if filters.get('start_date'):
        try:
            start_date = datetime.strptime(
                filters.get('start_date'), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            start_date = None

    if filters.get('end_date'):
        try:
            end_date = datetime.strptime(
                filters.get('end_date'), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            end_date = None

    summary_stats = get_liquidation_summary_stats(
        start_date=start_date,
        end_date=end_date,
        legislative_district=filters.get('legislative_district'),
        municipality=filters.get('municipality'),
        school_district=filters.get('school_district'),
        school_ids=filters.get('school_ids')
    )

    # Summary statistics in a compact table format
    summary_rows = [
        ('Total Liquidations:', summary_stats['total_liquidations']),
        ('Liquidated:', summary_stats['liquidated']),
        ('Pending Review:', summary_stats['pending_review']),
        ('Needs Revision:', summary_stats['needs_revision']),
        ('Draft:', summary_stats['draft'])
    ]

    # Create a compact 2-column layout for summary statistics
    for i, (label, value) in enumerate(summary_rows, start=summary_start_row + 1):
        # Label column
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{i}'].alignment = Alignment(
            horizontal='right', vertical='center')
        ws[f'A{i}'].border = thin_border

        # Value column
        ws[f'B{i}'] = value
        ws[f'B{i}'].font = Font(name='Calibri', size=11)
        ws[f'B{i}'].alignment = center_alignment
        ws[f'B{i}'].border = thin_border

    # ========== "PREPARED BY" SECTION ==========
    prepared_by_row = summary_start_row + \
        len(summary_rows) + 3  # +3 for spacing

    # Define light background fill for signature section
    light_bg_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    # Prepared by header
    ws[f'A{prepared_by_row}'] = "Prepared By:"
    ws[f'A{prepared_by_row}'].font = Font(name='Calibri', size=12)
    ws.merge_cells(f'A{prepared_by_row}:C{prepared_by_row}')
    # Center align the header and add background
    ws[f'A{prepared_by_row}'].alignment = Alignment(
        horizontal='center', vertical='center')
    ws[f'A{prepared_by_row}'].fill = light_bg_fill

    # Add accountant's e-signature if available (with proper spacing)
    if active_accountant and active_accountant.e_signature:
        try:
            signature_path = os.path.join(
                settings.MEDIA_ROOT, str(active_accountant.e_signature))
            if os.path.exists(signature_path):
                signature_img = Image(signature_path)
                signature_img.width = 120
                signature_img.height = 60
                # Position signature with more spacing from the text
                ws.add_image(signature_img, f'B{prepared_by_row + 2}')
        except Exception as e:
            logger.error(f"Could not add accountant signature: {e}")

    # Accountant name (use the active accountant)
    accountant_name = "Division Accountant"
    if active_accountant:
        accountant_name = f"{active_accountant.first_name} {active_accountant.last_name}".strip()
        if not accountant_name:
            accountant_name = active_accountant.email

    # Prepared by details - position name below signature
    ws[f'A{prepared_by_row + 5}'] = f"{accountant_name.upper()}"
    ws[f'A{prepared_by_row + 5}'].font = Font(
        name='Calibri', size=11, bold=True)
    ws.merge_cells(f'A{prepared_by_row + 5}:C{prepared_by_row + 5}')
    # Center align the name and add background
    ws[f'A{prepared_by_row + 5}'].alignment = Alignment(
        horizontal='center', vertical='center')
    ws[f'A{prepared_by_row + 5}'].fill = light_bg_fill

    ws[f'A{prepared_by_row + 6}'] = "Division Accountant"
    ws[f'A{prepared_by_row + 6}'].font = Font(name='Calibri', size=11)
    ws.merge_cells(f'A{prepared_by_row + 6}:C{prepared_by_row + 6}')
    ws[f'A{prepared_by_row + 6}'].alignment = Alignment(
        horizontal='center', vertical='center')
    ws[f'A{prepared_by_row + 6}'].fill = light_bg_fill

    # Add background to empty rows in signature section for visual consistency
    for row_offset in range(1, 6):  # Rows 1-5 (prepared_by_row + 1 to prepared_by_row + 5)
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{prepared_by_row + row_offset}']
            cell.fill = light_bg_fill

    # ws[f'A{prepared_by_row + 3}'] = f"Date Prepared: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    # ws[f'A{prepared_by_row + 3}'].font = Font(name='Calibri', size=11)
    # ws.merge_cells(f'A{prepared_by_row + 3}:D{prepared_by_row + 3}')

    # Signature line
    # ws[f'A{prepared_by_row + 5}'] = "_________________________"
    # ws[f'A{prepared_by_row + 5}'].font = Font(name='Calibri', size=11)
    # ws.merge_cells(f'A{prepared_by_row + 5}:D{prepared_by_row + 5}')

    # ws[f'A{prepared_by_row + 6}'] = "Signature over Printed Name"
    # ws[f'A{prepared_by_row + 6}'].font = Font(
    #     name='Calibri', size=10, italic=True)
    # ws.merge_cells(f'B{prepared_by_row + 6}:D{prepared_by_row + 6}')

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row_idx in range(1, data_start_row + len(report_data)):
            cell = ws.cell(row=row_idx, column=col_idx)
            is_merged = False
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    is_merged = True
                    break

            if not is_merged and cell.value:
                try:
                    cell_value_length = len(str(cell.value))
                    if cell_value_length > max_length:
                        max_length = cell_value_length
                except:
                    pass

        adjusted_width = min(max_length + 2, 50)
        if adjusted_width < 8:
            adjusted_width = 8
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Generate filename
    filename_parts = ['liquidation_report']
    if filters.get('start_date') and filters.get('end_date'):
        filename_parts.append(
            f"{filters['start_date']}_to_{filters['end_date']}")
    elif filters.get('status') and filters['status'] != 'all':
        filename_parts.append(filters['status'])

    response = HttpResponse(buffer.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{"_".join(filename_parts)}.xlsx"'

    return response

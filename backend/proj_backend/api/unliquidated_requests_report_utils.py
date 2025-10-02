# report_utils.py
import io
import csv
import os
from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone
from rest_framework.pagination import PageNumberPagination
from .models import RequestManagement
from django.db.models import Sum
from .models import RequestManagement
import logging
logger = logging.getLogger('api')


class AgingReportPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 1000


def get_aging_period(days):
    if days <= 30:
        return "0-30 days"
    elif days <= 60:
        return "31-60 days"
    elif days <= 90:
        return "61-90 days"
    elif days <= 120:
        return "91-120 days"
    elif days <= 180:
        return "121-180 days"
    else:
        return "180+ days"


def generate_aging_report_data(days_threshold='30'):
    """Generate aging report data for unliquidated requests"""
    # Get all unliquidated requests
    unliquidated_requests = RequestManagement.objects.filter(
        status='unliquidated')

    # Calculate aging for each request - use timezone-aware today
    today = timezone.now().date()  # This returns a date object, not datetime
    aging_data = []

    for req in unliquidated_requests:
        if req.downloaded_at:
            # downloaded_at is timezone-aware, convert both to date for comparison
            days_elapsed = (today - req.downloaded_at.date()).days
        else:
            # Fallback to created_at if downloaded_at is not available
            days_elapsed = (today - req.created_at.date()).days

        # Apply threshold filter
        should_include = False
        if days_threshold == 'all':
            should_include = True
        elif days_threshold == 'demand_letter':
            # Include requests that are exactly 29 days old (1 day before 30 days ends)
            should_include = days_elapsed == 29
        else:
            should_include = int(days_threshold) <= days_elapsed

        if should_include:
            aging_data.append({
                'request_id': req.request_id,
                'school_id': req.user.school.schoolId if req.user and req.user.school else '',
                'school_name': req.user.school.schoolName if req.user and req.user.school else '',
                'downloaded_at': req.downloaded_at.date() if req.downloaded_at else req.created_at.date(),
                'days_elapsed': days_elapsed,
                'aging_period': get_aging_period(days_elapsed),
                'amount': sum(rp.amount for rp in req.requestpriority_set.all())
            })

    return aging_data


def generate_aging_csv_report(aging_data, days_threshold):
    """Generate CSV report for aging data"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="aging_report_{days_threshold}_days.csv"'

    writer = csv.writer(response)

    # Add company header
    writer.writerow(['COMPANY NAME'])
    writer.writerow(['Aging Report'])
    writer.writerow([f'Generated on: {timezone.now().date()}'])
    writer.writerow([f'Days Threshold: {days_threshold}'])
    writer.writerow([])  # Empty row for spacing

    # Write column headers
    writer.writerow(['School ID', 'School Name', 'Request ID', 'Downloaded At',
                     'Days Elapsed', 'Aging Period', 'Amount'])

    for item in aging_data:
        writer.writerow([
            item['school_id'],
            item['school_name'],
            item['request_id'],
            item['downloaded_at'],
            item['days_elapsed'],
            item['aging_period'],
            item['amount']
        ])

    return response


def generate_aging_excel_report(aging_data, days_threshold, request_user=None):
    """Generate Excel report for aging data with same format as liquidation report"""
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Aging Report"

    # Define headers for the report
    headers = [
        'School ID', 'School Name', 'Request ID',
        'Downloaded At', 'Days Elapsed', 'Aging Period', 'Amount'
    ]

    # Add logo
    try:
        logo_path = os.path.join(
            settings.BASE_DIR, 'static', 'images', 'logo.png')
        if os.path.exists(logo_path):
            logo_img = Image(logo_path)
            logo_img.width = 110
            logo_img.height = 96
            ws.add_image(logo_img, 'B2')
    except Exception as e:
        logger.error(f"Could not add logo: {e}")

    # Define fonts and styles
    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=12, bold=True)
    data_font = Font(name='Calibri', size=11)
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    # Define border
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Add report header information
    ws['D3'] = "Department of Education"
    ws['D3'].font = title_font
    ws['D3'].alignment = center_alignment
    ws.merge_cells('D3:H3')

    ws['D4'] = "La Union Schools Division Office"
    ws['D4'].font = Font(name='Calibri', size=14, bold=True)
    ws['D4'].alignment = center_alignment
    ws.merge_cells('D4:H4')

    # Add title with threshold info
    report_title = "AGEING REQUESTS REPORT"
    threshold_text = "All requests" if days_threshold == 'all' else f"Requests older than {days_threshold} days"
    if days_threshold == 'demand_letter':
        threshold_text = "Requests at 29 days (Demand Letter)"

    ws.merge_cells('C5:J5')
    ws['C5'] = report_title
    ws['C5'].font = title_font
    ws['C5'].alignment = center_alignment

    ws['D6'] = f"Date Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['D6'].font = Font(name='Calibri', size=11)
    ws['D6'].alignment = center_alignment
    ws.merge_cells('D6:H6')

    # Add filter information
    filter_row = 8
    ws[f'A{filter_row}'] = "APPLIED FILTERS"
    ws[f'A{filter_row}'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{filter_row}:D{filter_row}')

    filter_info = []
    filter_info.append(f"Days Threshold: {threshold_text}")

    for i, filter_text in enumerate(filter_info, start=filter_row + 1):
        ws[f'A{i}'] = filter_text
        ws[f'A{i}'].font = Font(name='Calibri', size=10)
        ws.merge_cells(f'A{i}:D{i}')

    # Add column headers (starting after filters with proper spacing)
    header_row = filter_row + len(filter_info) + 2
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        # Add light gray background to header
        cell.fill = PatternFill(start_color="F0F0F0",
                                end_color="F0F0F0", fill_type="solid")

    # Set header row height
    ws.row_dimensions[header_row].height = 25

    # Add data rows
    data_start_row = header_row + 1
    for row_num, item in enumerate(aging_data, data_start_row):
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Set data row height
        ws.row_dimensions[row_num].height = 20

        # Fill in data
        ws.cell(row=row_num, column=1, value=item['school_id'])
        ws.cell(row=row_num, column=2, value=item['school_name'])
        ws.cell(row=row_num, column=3, value=item['request_id'])
        ws.cell(row=row_num, column=4, value=item['downloaded_at'])
        ws.cell(row=row_num, column=5, value=item['days_elapsed'])
        ws.cell(row=row_num, column=6, value=item['aging_period'])
        ws.cell(row=row_num, column=7, value=item['amount'])

    # ========== SUMMARY STATISTICS AT THE BOTTOM ==========
    summary_start_row = data_start_row + len(aging_data) + 3  # +3 for spacing

    # Summary header
    ws[f'A{summary_start_row}'] = "SUMMARY STATISTICS"
    ws[f'A{summary_start_row}'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells(f'A{summary_start_row}:D{summary_start_row}')

    # Calculate summary statistics
    total_requests = len(aging_data)

    # Count by aging period
    aging_periods = {
        "0-30 days": 0,
        "31-60 days": 0,
        "61-90 days": 0,
        "91-120 days": 0,
        "121-180 days": 0,
        "180+ days": 0
    }

    total_amount = sum(item['amount'] for item in aging_data)

    for item in aging_data:
        aging_periods[item['aging_period']] += 1

    # Summary statistics in a compact table format
    summary_rows = [
        ('Total Requests:', total_requests),
        ('Total Amount:', f"₱{total_amount:,.2f}"),
        ('', ''),
        ('Aging Breakdown:', ''),
        ('0-30 days:', aging_periods["0-30 days"]),
        ('31-60 days:', aging_periods["31-60 days"]),
        ('61-90 days:', aging_periods["61-90 days"]),
        ('91-120 days:', aging_periods["91-120 days"]),
        ('121-180 days:', aging_periods["121-180 days"]),
        ('180+ days:', aging_periods["180+ days"])
    ]

    # Create a compact 2-column layout for summary statistics
    for i, (label, value) in enumerate(summary_rows, start=summary_start_row + 1):
        # Label column
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(
            name='Calibri', size=11, bold=True) if i <= summary_start_row + 2 else Font(name='Calibri', size=11)
        ws[f'A{i}'].alignment = Alignment(
            horizontal='right', vertical='center')
        ws[f'A{i}'].border = thin_border

        # Value column
        ws[f'B{i}'] = value
        ws[f'B{i}'].font = Font(
            name='Calibri', size=11, bold=True) if i <= summary_start_row + 2 else Font(name='Calibri', size=11)
        ws[f'B{i}'].alignment = center_alignment
        ws[f'B{i}'].border = thin_border

    # ========== "PREPARED BY" SECTION ==========
    prepared_by_row = summary_start_row + \
        len(summary_rows) + 3  # +3 for spacing

    # Prepared by header
    ws[f'A{prepared_by_row}'] = "Prepared By:"
    ws[f'A{prepared_by_row}'].font = Font(name='Calibri', size=12)
    ws.merge_cells(f'A{prepared_by_row}:C{prepared_by_row}')
    # Center align the header
    ws[f'A{prepared_by_row}'].alignment = Alignment(
        horizontal='center', vertical='center')

    # Admin name (use the user who generated the report)
    admin_name = "System Administrator"
    if request_user:
        admin_name = f"{request_user.first_name} {request_user.last_name}".strip()
        if not admin_name:
            admin_name = request_user.email

    # Prepared by details
    ws[f'A{prepared_by_row + 1}'] = f"{admin_name.upper()}"
    ws[f'A{prepared_by_row + 1}'].font = Font(
        name='Calibri', size=11, bold=True)
    ws.merge_cells(f'A{prepared_by_row + 1}:C{prepared_by_row + 1}')
    # Center align the name
    ws[f'A{prepared_by_row + 1}'].alignment = Alignment(
        horizontal='center', vertical='center')

    ws[f'A{prepared_by_row + 2}'] = f"{request_user.role.title() if request_user else 'Administrator'}"
    ws[f'A{prepared_by_row + 2}'].font = Font(name='Calibri', size=11)
    ws.merge_cells(f'A{prepared_by_row + 2}:C{prepared_by_row + 2}')
    ws[f'A{prepared_by_row + 2}'].alignment = Alignment(
        horizontal='center', vertical='center')

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row_idx in range(1, data_start_row + len(aging_data)):
            cell = ws.cell(row=row_idx, column=col_idx)
            is_merged = False
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    is_merged = True
                    break

            if not is_merged and cell.value:
                try:
                    cell_value_length = len(str(cell.value))
                    if cell_value_length > max_length:
                        max_length = cell_value_length
                except:
                    pass

        adjusted_width = min(max_length + 2, 50)
        if adjusted_width < 8:
            adjusted_width = 8
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Generate filename based on threshold
    if days_threshold == 'demand_letter':
        filename = "aging_report_demand_letter.xlsx"
    else:
        filename = f"aging_report_{days_threshold}_days.xlsx"

    response = HttpResponse(buffer.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
